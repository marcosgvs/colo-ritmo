# -*- coding: utf-8 -*-
"""Gera: planilha do mês (grade+atendimento+pendências), PDF-calendário e .ics."""
import runpy, datetime as dt, os, zipfile
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ns = runpy.run_path("escala_out_v3.py")
PLAN, HOURS, METAS = ns["PLAN"], ns["HOURS"], ns["METAS"]
CM, CT, CN, mins, wd = ns["CM"], ns["CT"], ns["CN"], ns["mins"], ns["wd"]
cota_fds, CHUTES = ns["cota_fds"], ns["CHUTES"]

INK="3A2E2A"; INK2="6B5C56"; INK3="9A8A82"; LINE="EBE8E5"; CREME="FFFAF3"; CREME2="FAF3E8"
LAV="A299CB"; LAVI="5A4E8C"; LAVS="ECEAF4"; AQUA="9AD8E1"; AQUAI="6FA6CF"; AQUAS="E8F6F8"
SAND="E8C79A"; SANDS="FBF1E1"; CORAL="E7A59C"; CORALI="C77264"; CORALS="FBE9E5"
SAGE="A4D498"; SAGEI="5A6E50"; SAGES="ECF6E7"; BLUES="EAF2F9"; PINK="E79BC4"; OLIVES="F1EFE0"
F="Helvetica Neue"
FILLS={"M":SANDS,"T":BLUES,"D":SAND,"N":LAVI,"NT":LAV,"C":"FAEAF2","J":AQUAS,"E":OLIVES,
       "P":OLIVES,"R":OLIVES,"AB":PINK,"FE":AQUA,"LM":SAGE,"A":LINE}
WHITE={"N","NT"}
THIN=Side(style="thin",color=LINE); B=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
HDR=PatternFill("solid",fgColor=LAVI); HDRF=Font(name=F,color="FFFFFF",bold=True,size=10)

ORDER = ["Fred","Milena","Pabdo","Murilo","MSalomão","DebAlves","Vinicius","Amelio","Janaina",
 "Aline","CaAbreu","Danielle","Fabiula","Isabella","JuBrito","Kariny","Mayana","Neyde","Roberta","LuAlice",
 "Amanda","Fernando","João","JuCoutinho","LeLemos","Leomara","Marilia","Raylander","Ricardo","Rosana",
 "AnaSeverino","Anna","Ariadne","Beatriz","Bruna","Constantino","DebMatias","Denise","Ernesto","Grayce",
 "Heloa","Iggor","IsaRibeiro","Jaqueline","Joaquim","JuliaFig","Kozak","Laura","Leticia","LuCosta",
 "Marcia","MayWobido","Melara","Moabe","Nishioka","Patricia","Pedro","Pjamile","Raphael","Raquel",
 "Thamyres","Vanessa","Yuji","Stephanie","JuIsaac"]
WD_PT=["seg","ter","qua","qui","sex","sáb","dom"]

wb = openpyxl.Workbook(); wb.remove(wb.active)

# ---------------- GRADE OUT ----------------
ws = wb.create_sheet("OUT 2026"); ws.sheet_properties.tabColor = PINK; ws.sheet_view.showGridLines = False
c = ws.cell(row=1,column=1,value="outubro · 2026 — COMPLETA: mínimos 100% · 48 convocações por critério público · Laura e Moabe provisórios")
c.font = Font(name=F,bold=True,size=14,color=LAVI)
for i,d in enumerate(range(1,32)):
    col=3+i; L=get_column_letter(col); ws.column_dimensions[L].width=4.2
    fer = d==12
    for r,val in ((3,d),(4,WD_PT[wd(d)]),(5,"FER" if fer else None)):
        cc=ws.cell(row=r,column=col,value=val); cc.alignment=Alignment(horizontal="center")
        cc.font=Font(name=F,bold=True,size=9,color=CORALI if fer else (INK2 if wd(d)>=5 else INK))
        if fer: cc.fill=PatternFill("solid",fgColor=CORALS)
        elif wd(d)>=5: cc.fill=PatternFill("solid",fgColor=CREME2)
ws.cell(row=4,column=1,value="Nome").font=HDRF; ws.cell(row=4,column=1).fill=HDR
ws.cell(row=4,column=2,value="CH").font=HDRF; ws.cell(row=4,column=2).fill=HDR
ws.column_dimensions["A"].width=13; ws.column_dimensions["B"].width=4.5
sum0 = 3+31+1
for j,h in enumerate(["CH mês","Meta","Saldo","FDS h","Cota","SxN"]):
    cc=ws.cell(row=4,column=sum0+j,value=h); cc.fill=HDR; cc.font=HDRF
    cc.alignment=Alignment(horizontal="center"); ws.column_dimensions[get_column_letter(sum0+j)].width=7
r0=6
for i,ap in enumerate(ORDER):
    row=r0+i; p=PLAN.get(ap,{})
    nm=ws.cell(row=row,column=1,value=ap+(" *" if ap in CHUTES else "")); nm.font=Font(name=F,bold=True,size=10,color=INK)
    ch=ws.cell(row=row,column=2,value=METAS.get(ap,0)); ch.font=Font(name=F,size=9,color=INK3); ch.alignment=Alignment(horizontal="center")
    for k,d in enumerate(range(1,32)):
        col=3+k; cc=ws.cell(row=row,column=col); code=p.get(d)
        if code: cc.value=code
        cc.alignment=Alignment(horizontal="center"); cc.border=B; cc.font=Font(name=F,size=9,color=INK)
        if d==12: cc.fill=PatternFill("solid",fgColor=CORALS)
        elif wd(d)>=5: cc.fill=PatternFill("solid",fgColor=CREME2)
    tot=sum(HOURS.get(c2,0) for c2 in p.values())
    fe=sum(1 for c2 in p.values() if c2 in ("FE","LM"))
    meta=round(METAS.get(ap,0)/7*(31-fe),1)
    fds=sum(HOURS.get(c2,0) for d,c2 in p.items() if wd(d)>=5 and c2 not in ("FE","LM","AB"))
    sxn=sum(1 for d,c2 in p.items() if wd(d)==4 and c2=="N")
    vals=[tot,meta,round(tot-meta,1),fds,cota_fds(ap) if METAS.get(ap) else 0,sxn]
    for j,v in enumerate(vals):
        cc=ws.cell(row=row,column=sum0+j,value=v); cc.font=Font(name=F,size=8.5,color=INK2); cc.alignment=Alignment(horizontal="center")
nrows=len(ORDER)
cov0=r0+nrows+1
for li,(lab,codes,mi) in enumerate([("Manhã",CM,0),("Tarde",CT,1),("Noite",CN,2)]):
    row=cov0+li
    ws.cell(row=row,column=1,value="◆ "+lab).font=Font(name=F,bold=True,size=9,color=LAVI)
    for k,d in enumerate(range(1,32)):
        col=3+k
        n=sum(1 for ap2 in ORDER if PLAN.get(ap2,{}).get(d) in codes)
        mn=mins(d)[mi]
        cc=ws.cell(row=row,column=col,value=n); cc.alignment=Alignment(horizontal="center")
        ok = n>=mn
        cc.font=Font(name=F,size=9,bold=True,color=SAGEI if ok else "FFFFFF")
        if not ok: cc.fill=PatternFill("solid",fgColor=CORAL)
ws.cell(row=cov0+3,column=1,value="mínimos: úteis 14/10/7 · sáb 10/8/7 · dom e feriado 9/8/7 · * = sem preferências (padrão set)").font=Font(name=F,size=8,italic=True,color=INK3)
rng_=f"C{r0}:{get_column_letter(33)}{r0+nrows-1}"
for code,color in FILLS.items():
    fnt=Font(name=F,color="FFFFFF",bold=True) if code in WHITE else Font(name=F,color=INK)
    ws.conditional_formatting.add(rng_,CellIsRule(operator="equal",formula=[f'"{code}"'],fill=PatternFill("solid",fgColor=color),font=fnt))
ws.freeze_panes="C6"

# ---------------- ATENDIMENTO ----------------
ATEND = [
 ("Sexta-noite 16/10","CRITÉRIO PÚBLICO","10 candidatos para 7 vagas. Desempate: menos sextas-noite em 2026. Concedidas: Jaqueline (0 no ano), Marília (3 · caso autorizado pela Paula), Bruna (3), DebMatias (3), JuCoutinho (3), Roberta (3), Mayana (4). Não couberam: JuliaFig (4 → recebeu a sexta 30, alternativa que ela mesma ofereceu), Grayce (5), MayWobido (5), IsaRibeiro (7, maior contagem do grupo)."),
 ("Kozak","✓ concessões","Folga no sáb 31/10 (aniversário da filha) CONCEDIDA — cobre 29 e 30/10 noite como ela propôs. Férias 05–19 aplicadas. Sexta 2/10 dela usada."),
 ("Murilo","✓ 100%","Sábado NOTURNO 15/15 concedido (10 e 24) + rotina onco nas manhãs."),
 ("MayWobido","◐ parcial","Abono aniversário 31/10 CONCEDIDO ✓. Sexta 16 negada (critério acima) — ficou com 9N, que ela ofereceu."),
 ("Grayce","◐ confirmar","Sexta 16 negada (5 SN no ano). Reposição sugerida: sexta 9/10 noite (FORA das preferências dela — CONFIRMAR antes de publicar). CH fica -10h se recusar."),
 ("IsaRibeiro","◐ parcial","Sexta 16 negada (7 SN no ano). Recebeu 1/10 N e 17/10 N fora das preferências: eram os dias com menos elegíveis do mês e ela é a única sem veto — reconhecer o crédito em novembro."),
 ("Danielle","◐ parcial","17 e 22/10 N fora das preferências (sem veto dela; buracos das férias da Kozak). Resto 100% conforme e-mail."),
 ("Constantino","◐ parcial","4/10 N fora das preferências (domingo deficitário; ele só vetou 10-11 e o feriado). Feriado 12 e fds 10-11 folgados como pediu ✓."),
 ("Anna","◐ leve","4/10 TARDE fora das preferências (ela só vetou domingo NOITE). Sexta 30 preferida dela ✓ + 31 N ✓."),
 ("MSalomão","⚠ validar","O plano DELA contém 18h invertido 2×: 4N→5M e 18N→19M. Mantido porque foi pedido explícito — Mari decide se aceita."),
 ("Aline","⚠ validar","Plano dela: noite de 5/10 emenda com reunião do CEP às 8h de 6/10. Mantido (pedido dela) — validar. FDS 18h vs cota 24h (plano dela)."),
 ("JuBrito","✓ com nota","Folga no feriado ✓. CH 114h vs meta 133h (-19h): redução para 30h + poucas ofertas — vira crédito/banco para novembro."),
 ("JuliaFig","✓ com nota","-22h no mês: ofertou 7 noites e 2 caem em novembro (01-02/11). Banco a favor dela."),
 ("Amelio","✓ 100%","Fora de Brasília 24-31 com BHN, compensado com BHP em 3 e 16/10 — exatamente o desenho dela."),
 ("Thamyres","✓ com nota","Feriado 12 (7-19h) como ela ofereceu ✓. FDS 12h vs cota 18h: viagem 13-18 + a sugestão dela de 01/11 conta em novembro."),
 ("Rosana","✓ 100%","Sexta 23 usada (opção dela) · fds 17-18 dela ✓ · dia 13 protegido ✓."),
 ("LuCosta","✓ 100%","O 9/10 N 'se precisar' foi usado — precisamos (sexta com 5 vetos)."),
 ("Fabiula","✓ com nota","FDS 24h vs cota 36h: padrão dela de 1 fds/mês; compensar em novembro."),
 ("48 convocações","critério público","Cada convocação escolhida por: 1º menor saldo de horas no mês, 2º menos convocações. Lista completa na aba CONVOCAÇÕES. Todas geram crédito no placar de novembro."),
 ("Laura + Moabe","* provisório","Sem preferências — padrão de setembro. Refazer as linhas quando enviarem."),
 ("Ariadne ∥ Raylander","✓ regra do casal","Nenhum choque de turno na semana do retorno dela (26-31) ✓."),
]
ws2 = wb.create_sheet("ATENDIMENTO"); ws2.sheet_properties.tabColor = PINK; ws2.sheet_view.showGridLines=False
c=ws2.cell(row=1,column=1,value="atendimento e justificativas · outubro 2026"); c.font=Font(name=F,bold=True,size=14,color=LAVI)
c=ws2.cell(row=2,column=1,value="Cobertura 100% nos 31 dias · 40 de 58 com 100% das preferências · convocações na aba CONVOCAÇÕES · 3 itens para validação da Mari")
c.font=Font(name=F,size=10,italic=True,color=INK2)
for j,(h,w) in enumerate(zip(["Quem / o quê","Status","Justificativa (critério público)"],[22,16,120]),1):
    cc=ws2.cell(row=4,column=j,value=h); cc.fill=HDR; cc.font=HDRF; ws2.column_dimensions[get_column_letter(j)].width=w
for i,(quem,st,txt) in enumerate(ATEND,5):
    ws2.cell(row=i,column=1,value=quem).font=Font(name=F,bold=True,size=10,color=INK)
    s=ws2.cell(row=i,column=2,value=st); s.font=Font(name=F,bold=True,size=9,color=SAGEI if st.startswith("✓") else CORALI)
    t=ws2.cell(row=i,column=3,value=txt); t.font=Font(name=F,size=9,color=INK); t.alignment=Alignment(wrap_text=True,vertical="top")
    for j in (1,2,3): ws2.cell(row=i,column=j).border=B
ws2.freeze_panes="A5"

# ---------------- PENDÊNCIAS ----------------
ns2 = runpy.run_path("escala_out_v3.py")
CONVOC = ns2["CONVOC"]
conv_txt = " · ".join(f"{d:02d}/{['seg','ter','qua','qui','sex','sáb','dom'][ns2['wd'](d)]} {s}: {ap}" for ap, d, s in sorted(CONVOC, key=lambda x: (x[1], x[2])))
PEND = [
 ("COMO A ESCALA FECHOU","Objetivo 1 cumprido: mínimos 100% nos 31 dias (14/10/7 útil · sáb 10/8/7 · dom/feriado 9/8/7). Buracos preenchidos com 48 convocações fora das preferências, escolhidas por critério público: 1º quem estava abaixo da própria carga no mês, 2º quem tinha menos convocações. Impedimentos duros (motivo declarado: outro serviço, filhos, viagem, atestado) foram INTOCADOS — ninguém foi convocado contra impedimento real."),
 ("CONVOCAÇÕES (48)", conv_txt),
 ("Mais convocados","Ricardo (4) · Pedro, Bruna, DebMatias, Jaqueline, MayWobido, LuCosta, Moabe*, Leticia, Roberta, Patricia, Nishioka, JuBrito (2-3 cada). Todos entram com CRÉDITO no placar de novembro — prioridade nas preferências."),
 ("Horas a mais (pagas)","As convocações geram +12 a +21h acima da carga em ~12 pessoas — padrão normal do grupo (a contagem anual sempre teve 'horas a mais'). Quem ficou DEVENDO: JuliaFig -22h (2 noites dela caem em 01-02/11), Amelio -21h (BHN desenhado por ela), Ernesto e Grayce -10h (grades rígidas) — vira banco."),
 ("Validar com as pessoas","1) João 22/10 qui NOITE — ele vetou qui noite na semana 3, não na 4; confirmar. 2) Ricardo pediu 'avisar antes de mudar' — avisá-lo das 4 convocações. 3) Grayce: sexta 9 N como reposição da 16 negada. 4) MSalomão e Aline: emendas de 18h nos planos delas. 5) Laura e Moabe: linhas provisórias (sem preferências)."),
]
ws3 = wb.create_sheet("CONVOCAÇÕES"); ws3.sheet_properties.tabColor = CORALI; ws3.sheet_view.showGridLines=False
c=ws3.cell(row=1,column=1,value="convocações e validações · outubro 2026"); c.font=Font(name=F,bold=True,size=14,color=CORALI)
for j,(h,w) in enumerate(zip(["Tema","Detalhe / decisão necessária"],[26,130]),1):
    cc=ws3.cell(row=3,column=j,value=h); cc.fill=HDR; cc.font=HDRF; ws3.column_dimensions[get_column_letter(j)].width=w
for i,(tema,txt) in enumerate(PEND,4):
    ws3.cell(row=i,column=1,value=tema).font=Font(name=F,bold=True,size=10,color=INK)
    t=ws3.cell(row=i,column=2,value=txt); t.font=Font(name=F,size=9,color=INK); t.alignment=Alignment(wrap_text=True,vertical="top")
    for j in (1,2): ws3.cell(row=i,column=j).border=B
wb.save("Escala UTI HCB - outubro 2026.xlsx")
print("xlsx ok")

# ---------------- ICS por médico ----------------
SHIFT_TIMES = {"M":("070000","130000"),"T":("130000","190000"),"D":("070000","190000"),
               "N":("190000","070000+1"),"NT":("190000","010000+1"),"C":("080000","190000"),
               "J":("080000","130000"),"E":("080000","120000"),"P":("070000","130000"),"R":("130000","190000")}
NAMES = {"M":"Manhã","T":"Tarde","D":"Dia 12h","N":"Noite","NT":"Noitinha","C":"Chefia 10h","J":"Manhã (8-13h)","E":"CEP","P":"Paliativo","R":"CRO"}
os.makedirs("ics", exist_ok=True)
zf = zipfile.ZipFile("Escala outubro 2026 - calendarios individuais.zip","w",zipfile.ZIP_DEFLATED)
for ap in ORDER:
    p = PLAN.get(ap,{})
    evs=[]
    for d in sorted(p):
        code=p[d]
        if code not in SHIFT_TIMES: continue
        ini,fim=SHIFT_TIMES[code]
        d2 = d+1 if fim.endswith("+1") else d
        fim=fim.replace("+1","")
        evs.append(
            "BEGIN:VEVENT\r\nUID:hcb-out26-%s-%d@coloritmo\r\nDTSTART;TZID=America/Sao_Paulo:202610%02dT%s\r\nDTEND;TZID=America/Sao_Paulo:202610%02dT%s\r\nSUMMARY:UTI HCB · %s\r\nEND:VEVENT" % (ap,d,d,ini,min(d2,31),fim,NAMES[code]))
    if not evs: continue
    ics="BEGIN:VCALENDAR\r\nVERSION:2.0\r\nPRODID:-//colo ritmo//escala HCB//PT\r\nCALSCALE:GREGORIAN\r\n"+"\r\n".join(evs)+"\r\nEND:VCALENDAR\r\n"
    zf.writestr(f"{ap} - outubro 2026.ics", ics)
zf.close()
print("ics zip ok")
