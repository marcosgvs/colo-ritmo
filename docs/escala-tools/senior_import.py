#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Importa o histórico real da escala UTI HCB a partir dos arquivos de códigos Senior.

Fonte: ~/Downloads/fwdarquivosdaescala/ESCALA CÓDIGOS ... .xlsx (fev, mar, abr, jul, ago, set/2026).
Saída: DIAS[date][apelido] = (letra, codigo_senior)

Cuidados que este módulo resolve:
- nomes de arquivo em NFD no macOS (normalizar antes de casar acento);
- cabeçalho em linha 1 OU 2 dependendo do arquivo;
- a grade corre por SEMANAS COMPLETAS: há colunas de dias do mês seguinte
  (março vai até 05/04, julho até 02/08). Quando dois arquivos cobrem a mesma
  data, vence o arquivo cujo mês nominal é o mês da data;
- códigos antigos (jan–abr) 110/82/83 e novos (mai em diante) 2/40/41;
- mapa nome-completo -> apelido EXPLÍCITO (nunca fuzzy: BHP/BHN já enganaram
  fuzzy match antes).
"""
import datetime as dt
import glob
import os
import re
import unicodedata

import openpyxl

DIR_FONTES = os.path.expanduser("~/Downloads/fwdarquivosdaescala")

nfc = lambda s: unicodedata.normalize("NFC", s)

MESES = {"JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
         "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12}

# ---------------------------------------------------------------- códigos Senior
# letra interna, horas, janela (início, fim em horas decimais; fim > 24 = vira o dia)
CODIGOS = {
    # pós-mudança de abril
    "2":    ("M",  6.0,  7.0, 13.0),
    "239":  ("T",  6.0, 13.0, 19.0),
    "40":   ("D", 12.0,  7.0, 19.0),
    "41":   ("N", 12.0, 19.0, 31.0),
    # pré-abril (jan–abr)
    "110":  ("M",  6.0,  7.0, 13.0),
    "82":   ("D", 12.0,  7.0, 19.0),
    "83":   ("N", 12.0, 19.0, 31.0),
    # estruturais
    "47":   ("C", 10.0,  8.0, 19.0),   # 10h chefia (8–12 + 13–19)
    "11":   ("A",  8.0,  8.0, 17.0),   # 8h administrativo
    "78":   ("J",  5.0,  8.0, 13.0),   # 5h Janaina
    "10":   ("J",  5.0,  8.0, 13.0),   # idem, código antigo
    "6":    ("CEP", 4.0, 8.0, 12.0),  # 4h CEP
    "9":    ("t",  4.0, 13.0, 17.0),   # 4h tarde
    "15":   ("q",  5.0, 13.0, 18.0),   # 5h tarde
    "367":  ("z", 10.0,  8.0, 18.0),   # 10h 8–18
    "349":  ("NT", 6.0, 19.0, 25.0),   # 6h noitinha (vestigial)
    # combinações de 18h — existem na legenda, proibidas na prática
    "1000": ("X", 18.0, 13.0, 31.0),   # tarde + noite
    "364":  ("X", 18.0,  7.0, 31.0),   # manhã + noite
    "365":  ("X", 18.0,  7.0, 25.0),   # dia + noitinha
    "373":  ("X", 12.0, 13.0, 25.0),   # tarde + noitinha
}
AUSENCIAS = {"FÉRIAS": "FE", "FERIAS": "FE", "LM": "LM", "LICENÇA": "LM", "ATESTADO": "AT"}

# ------------------------------------------------------- nome completo -> apelido
MAPA_NOMES = {
    "Aline Saliba": "Aline",           "Amanda Braga": "Amanda",
    "Ana Kozak": "Kozak",              "Ana Severino": "AnaSeverino",
    "Andréa Kairala": "Kairala",       "Anna": "Anna",
    "Anna Jorge": "Anna",              "Ariadne": "Ariadne",
    "Beatriz": "Beatriz",              "Bruna": "Bruna",
    "Camila Abreu": "CaAbreu",         "Carlos Ernesto": "Ernesto",
    "Daniel Raylander": "Raylander",   "Danielle Tanajura": "Danielle",
    "Dayana": "Dayana",                "Debora Matias": "DebMatias",
    "Deborah Alves": "DebAlves",       "Denise": "Denise",
    "Ernesto": "Ernesto",              "Fabiula Czameski": "Fabiula",
    "Fernanda Amelio": "Amelio",       "Fernanda Constantino": "Constantino",
    "Fernanda Kariny": "Kariny",       "Fernando Filardi": "Fernando",
    "Frederico Pires": "Fred",         "Grayce Maya": "Grayce",
    "Heloa": "Heloa",                  "Henrique Yuji": "Yuji",
    "Iggor Almeida": "Iggor",          "Isabela Ribeiro": "IsaRibeiro",
    "Isabella Mazzaro": "Isabella",    "Janaina Rabelo": "Janaina",
    "Jaqueline": "Jaqueline",          "Joaquim": "Joaquim",
    "João": "João",                    "Julia Figueiredo": "JuliaFig",
    "Julia Isaac": "JuIsaac",          "Juliana Coutinho": "JuCoutinho",
    "Julliana Brito": "JuBrito",       "Laura Haydee": "Laura",
    "Laura Nishioka": "Nishioka",      "Leomara": "Leomara",
    "Leticia Lemos": "LeLemos",        "Letícia Café": "Leticia",
    "Luciana Alice": "LuAlice",        "Luciana Costa": "LuCosta",
    "Luciana Melara": "Melara",        "Marcia": "Marcia",
    "Mariana Pinheiro": "MPinheiro",   "Marilia": "Marilia",
    "Marina Salomão": "MSalomão",      "Mayana Leal": "Mayana",
    "Mayara Wobido": "MayWobido",      "Milena": "Milena",
    "Moabe": "Moabe",                  "Murilo": "Murilo",
    "Neyde Brito": "Neyde",            "Patricia": "Patricia",
    "Paula Abdo": "Pabdo",             "Paula Jamile": "Pjamile",
    "Pedro": "Pedro",                  "Raphael Costa": "Raphael",
    "Raquel Assis": "Raquel",          "Ricardo": "Ricardo",
    "Roberta Iglesias": "Roberta",     "Rosana": "Rosana",
    "Stephanie": "Stephanie",          "Thamyres": "Thamyres",
    "Vanessa": "Vanessa",              "Vinicius Bezerra": "Vinicius",
}

PARA_RODAPE = ("LEGENDA", "FÉRIAS", "FERIAS", "TOTAL", "SOMA", "COBERTURA",
               "ANIVERS", "LICENÇA", "OBS", "MANHÃ", "TARDE", "NOITE")


def _arquivos_codigos():
    saida = []
    for f in sorted(nfc(x) for x in glob.glob(os.path.join(DIR_FONTES, "*.xlsx"))):
        base = os.path.basename(f).upper()
        if "DIGOS" not in base or "SESMT" in base or "(1)" in base:
            continue
        mes = next((n for k, n in MESES.items() if k in base), None)
        if mes:
            saida.append((mes, f))
    return sorted(saida)


def _linha_cabecalho(ws):
    for r in range(1, 10):
        v = ws.cell(row=r, column=1).value
        if v and "DATA" in str(v).upper():
            return r
    raise ValueError("cabeçalho 'DATA MÊS' não encontrado")


def ler_mes(caminho):
    """Devolve (dias_por_pessoa, orfaos) de um arquivo de códigos."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = _linha_cabecalho(ws)
    colunas = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        if isinstance(v, (dt.datetime, dt.date)):
            colunas[c] = v.date() if isinstance(v, dt.datetime) else v
    dados, orfaos = {}, []
    for r in range(hdr + 2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        up = s.upper()
        if up.startswith(PARA_RODAPE) or re.match(r"^\d+\s*[-–]", s):
            break
        if ":" in s or re.search(r"\d\d/\d\d", s):
            break
        apelido = MAPA_NOMES.get(nfc(s))
        if apelido is None:
            orfaos.append(s)
            continue
        for c, data in colunas.items():
            cel = ws.cell(row=r, column=c).value
            if cel in (None, ""):
                continue
            bruto = str(cel).strip()
            if not bruto:
                continue
            dados.setdefault(apelido, {})[data] = bruto
    return dados, orfaos


def traduzir(bruto):
    """código Senior bruto -> (letra, codigo) ; letra '?' se desconhecido."""
    s = nfc(str(bruto).strip())
    chave = s.upper().rstrip(".")
    for k, v in AUSENCIAS.items():
        if chave.startswith(k):
            return v, s
    s_num = s.split(".")[0].split()[0] if s else s
    if s_num in CODIGOS:
        return CODIGOS[s_num][0], s_num
    return "?", s


def importar():
    """DIAS[date][apelido] = (letra, codigo) + relatório de conflitos/órfãos."""
    DIAS, conflitos, orfaos, desconhecidos = {}, [], set(), {}
    for mes, caminho in _arquivos_codigos():
        dados, orf = ler_mes(caminho)
        orfaos.update(orf)
        for apelido, por_data in dados.items():
            for data, bruto in por_data.items():
                letra, codigo = traduzir(bruto)
                if letra == "?":
                    desconhecidos.setdefault(codigo, []).append((apelido, data))
                    continue
                atual = DIAS.setdefault(data, {})
                novo = (letra, codigo)
                if apelido in atual and atual[apelido] != novo:
                    # quem manda é o arquivo cujo mês nominal == mês da data
                    vencedor = novo if mes == data.month else atual[apelido]
                    conflitos.append((data, apelido, atual[apelido], novo, vencedor))
                    atual[apelido] = vencedor
                elif apelido not in atual:
                    atual[apelido] = novo
                # se o arquivo do mês certo já escreveu, um arquivo vizinho não sobrescreve
    return DIAS, {"conflitos": conflitos, "orfaos": sorted(orfaos),
                  "desconhecidos": desconhecidos}


if __name__ == "__main__":
    DIAS, rel = importar()
    datas = sorted(DIAS)
    print(f"datas cobertas: {len(datas)}  ({datas[0]} → {datas[-1]})")
    por_mes = {}
    for d in datas:
        por_mes.setdefault(d.month, 0)
        por_mes[d.month] += 1
    print("dias por mês:", dict(sorted(por_mes.items())))
    lanc = sum(len(v) for v in DIAS.values())
    print(f"lançamentos: {lanc}")
    print(f"órfãos de nome: {rel['orfaos'] or 'nenhum'}")
    print(f"códigos desconhecidos: { {k: len(v) for k, v in rel['desconhecidos'].items()} or 'nenhum'}")
    for k, v in rel["desconhecidos"].items():
        print(f"   código {k!r}: {v}")
    print(f"conflitos entre arquivos: {len(rel['conflitos'])}")
    for c in rel["conflitos"][:15]:
        print("   ", c)
