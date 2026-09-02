#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Planilha V3 — ANUAL, UTI HCB 2026 (gerador; a numeração V3 é a das versões
que a Mari recebe).

Abas: DASHBOARD · LEIA-ME · CADASTRO · CONFIG · 12 mensais (JAN..DEZ) · OUT · DIA A DIA
      · CHECAGEM OUT · PAINEL ANO · SENIOR · VALIDADOR · DADOS DASH

Só realidade: histórico dos arquivos oficiais (Senior + grades do grupo), outubro
na versão viva da Mari + as correções da checagem dela e do Marcos. Nada de
proposta do gerador (as abas ATENDIMENTO/CONVOCAÇÕES/ESBOÇO saíram em 01/09/26).

Princípios:
- regras vivem em CONFIG como DADO; as abas mensais apontam pra lá (mudou o
  mínimo em CONFIG, a planilha inteira se ajusta);
- todas as 12 abas têm GEOMETRIA IDÊNTICA, o que deixa SENIOR e PAINEL ANO
  espelharem qualquer mês por INDIRECT sem gambiarra;
- contagem é fórmula, nunca número digitado — o erro de contagem à mão foi o
  que gerou as dívidas de fds de agosto;
- compatível com Google Sheets: só COUNTIF/SUMPRODUCT/VLOOKUP/INDIRECT/ADDRESS,
  e formatação condicional só nas 3 linhas de "falta" (CF pesada engasga o Sheets).
"""
import datetime as dt
import os
import runpy
import sys

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

AQUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, AQUI)

# MODO "v3": planilha nova (grades como fonte, correções da checagem, aba CHECAGEM).
# MODO "antiga": o Sheet que a Mari usa (id em SHEET_ANTIGO) recebe SÓ a estrutura —
# Sem/BH, cabeçalho vertical, códigos novos, alturas — com os dados como estão lá:
# outubro lido AO VIVO na hora de gerar, jan–set pelas fontes de antes (Senior +
# grades de jan/mai/jun), nenhuma correção da checagem. Pedido do Marcos, 02/09/26.
MODO = "antiga" if "antiga" in sys.argv[1:] else "v3"
SHEET_ANTIGO = "102d4E3IzlSXH4MDU6hd9ywr21BxthHdqL9meBP-IX7s"
RENOMEADOS = {"E": "CEP", "P": "CP", "R": "CRO"}   # códigos que ganharam nome legível

import grade_import
import senior_import
import v4_dados as D
from validador import auditar, noturnas_por_mes

# ------------------------------------------------------------- identidade colo
INK, INK2, INK3 = "3A2E2A", "6B5C56", "9A8A82"
LINE, LINE2 = "EBE8E5", "DAD3CD"
CREME, CREME2 = "FFFAF3", "FAF3E8"   # bg e surface dos tokens
LAV, LAVI, LAVS = "A299CB", "5A4E8C", "ECEAF4"
AQUA, AQUAS = "9AD8E1", "E8F6F8"
SAND, SANDS = "E8C79A", "FBF1E1"
CORAL, CORALI, CORALS = "E7A59C", "C77264", "FBE9E5"
SAGE, SAGES = "A4D498", "ECF6E7"
BLUES, PINK, OLIVES = "EAF2F9", "E79BC4", "F1EFE0"
F = "Nunito"          # corpo e UI
DISPLAY = "Fraunces"  # títulos
MAO = "Caveat"        # sobretítulo à mão

FILLS = {"M": SANDS, "T": BLUES, "D": SAND, "N": LAVI, "NT": LAV, "C": "FAEAF2",
         "J": AQUAS, "CEP": OLIVES, "CP": "E9F0E4", "CRO": "EDE7F6", "A": LINE,
         "FE": AQUA, "LM": SAGE, "AB": PINK}
# banco de horas: toda célula BHP (+) pinta de coral claro, toda BHN (−) de cinza
# riscado — a cor diz o estado do banco, a letra diz o turno
FILL_BHP, FILL_BHN = CORALS, LINE
BRANCO = {"N"}
THIN = Side(style="thin", color=LINE)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WD_PT = ["seg", "ter", "qua", "qui", "sex", "sáb", "dom"]
MESES_PT = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN",
            "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
MESES_LONGO = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho",
               "agosto", "setembro", "outubro", "novembro", "dezembro"]

# ------------------------------------------------------------------ geometria
C_MED, C_CH = 1, 2
N_PRE = 6                     # colunas do FIM DO MÊS ANTERIOR: todo mês abre com a
                              # semana civil fechada (dia 1º cai terça → a segunda
                              # aparece). São fórmulas vivas da aba anterior — entram
                              # nas somas semanais, mas CH/FDS/saldo delas pertencem
                              # ao mês anterior (espelho da regra do dia 1º seguinte)
C_PRE1 = 3                    # primeira coluna de véspera (C)
C_D1 = C_PRE1 + N_PRE         # primeira coluna de dia (I)
C_DN = C_D1 + 30              # último dia DO MÊS (AM) — CH/FDS/saldo param aqui
N_POS = 6                     # colunas da VIRADA: a semana fecha NO DOMINGO
                              # (Marcos, 28/08), então o mês mostra os dias do mês
                              # seguinte até o domingo que fecha a última semana
                              # (nov termina seg 30/11 → aparecem 01–06/12).
                              # Fórmulas vivas da aba seguinte; contam nas somas
                              # SEMANAIS, mas a cota é do mês seguinte (doc §6)
C_D32 = C_D1 + 31             # primeira coluna da virada (AN)
N_SEM = 6                     # máximo de semanas civis que um mês encosta
# As semanas vêm PRIMEIRO (pedido do Marcos, 01/09/26): Sem k = horas da semana
# civil, BH k = Sem k − alvo semanal da pessoa (positivo BHP, negativo BHN).
COLS_SEM = [x for k in range(1, N_SEM + 1) for x in (f"Sem {k}", f"BH {k}")]
COLS_TOT = (COLS_SEM
            + ["CH mês", "FDS", "SxN", "Feriado", "Meta", "Saldo", "18h⚠", "N→T",
               "Cota FDS", "FDS⚠", "Sem⚠"]      # fds⚠ = excesso sobre cota × fator
            + ["Grupo",       # ordenar por tipo no funil do filtro
               "Nº"])         # posição original do cadastro — ordena de volta
IDX = {nome: i for i, nome in enumerate(COLS_TOT)}
COLS_ALERTA = {"18h⚠", "N→T", "Cota FDS", "FDS⚠", "Sem⚠"}
IDX_GRUPO = IDX["Grupo"]
IDX_NUM = IDX["Nº"]


def col_tot(nome):
    """coluna (1-based) de um total nas abas mensais, pelo NOME — nunca por offset."""
    return C_TOT + IDX[nome]
# Grupo com prefixo numérico pra ordenar coordenação→rotina→staff→administrativo
# (o sortSpec do Sheets só sabe asc/desc). Coordenação = quem faz o 47 (doc §3).
GRUPO_COORD = {"Fred", "Milena"}
GRUPO_ROSTER = {a: g for a, _n, _c, g, *_ in D.ROSTER}
def grupo_filtro(apelido):
    if apelido in GRUPO_COORD:
        return "1 · coordenação"
    g = GRUPO_ROSTER.get(apelido, "")
    if g in ("chefia", "rotina"):
        return "2 · rotina"
    if g == "administrativo":
        return "4 · administrativo"
    return "3 · staff"
# Sem 1..6 = horas por SEMANA CIVIL (seg–dom), sempre COMPLETA: a 1ª inclui os
# dias do mês anterior e a última fecha no domingo com os dias do mês seguinte.
# A semana da virada aparece nos dois meses — é a mesma semana, medida igual.
# sem⚠ = a maior delas. Teto duro: 44h (art. 7º XIII da Constituição).
C_TOT = C_D32 + N_POS         # AT em diante
R_TIT, R_DIA, R_DOW, R_FDS, R_SEX, R_FER, R_HDR, R_P0 = 1, 2, 3, 4, 5, 6, 7, 8

L_PRE1 = get_column_letter(C_PRE1)
L_PRE2 = get_column_letter(C_PRE1 + 1)
L_D1, L_DN = get_column_letter(C_D1), get_column_letter(C_DN)
L_DN_1 = get_column_letter(C_DN - 1)
L_D2 = get_column_letter(C_D1 + 1)


def _ref_vizinho(aba, dia, lin):
    """célula do dia `dia` da pessoa desta linha na aba vizinha, PELO NOME.

    INDEX/MATCH com linhas absolutas: sobrevive à ordenação pelo filtro em
    qualquer uma das duas abas (a lição do #REF de 28/08 — referência posicional
    quebra quando as linhas se movem). O &"" evita o 0 de célula vazia.
    """
    colv = get_column_letter(C_D1 + dia - 1)
    return (f'=IFERROR(INDEX({aba}!{colv}${R_P0}:{colv}${R_P0 + 65},'
            f'MATCH($A{lin},{aba}!$A${R_P0}:$A${R_P0 + 65},0))&"","")')


def janelas_semana_civil(mes, ndias, ano=2026):
    """[[(col_ini, col_fim), …]] — cada semana civil do mês como lista de
    trechos contíguos de coluna. Toda semana é COMPLETA (segunda a domingo):
    a 1ª começa nas vésperas (C..H) e a última fecha no domingo, entrando
    pelas colunas da virada. Em mês de menos de 31 dias existe um vão de
    colunas cinzas entre o último dia e o bloco da virada — daí os trechos.
    """
    offset = dt.date(ano, mes, 1).weekday()          # seg=0 … dom=6
    janelas = []
    ini = -offset                                    # posição 0 = dia 1º
    while ini < ndias:
        fim = ini + 6
        spans = [(C_D1 + ini, C_D1 + min(fim, ndias - 1))]
        if fim >= ndias:                             # transborda pra virada
            spans.append((C_D32, C_D32 + (fim - ndias)))
        janelas.append(spans)
        ini += 7
    return janelas

# expressão de horas de uma linha de pessoa (usada em várias fórmulas).
# BHN puro (M- T- D- N-) tem 0h na tabela e por isso não entra; Tm-/Mt- entram com 6h
def expr_horas(lin, c1=None, c2=None):
    a = f"${get_column_letter(c1 or C_D1)}{lin}:${get_column_letter(c2 or C_DN)}{lin}"
    partes = [f'({a}="{k}")*{v[2]}' for k, v in D.TURNOS.items() if v[2]]
    return "+".join(partes)


def expr_conta(faixa, letras):
    """(faixa=letra)+(faixa=letra)… — 1 onde a célula é um dos códigos."""
    return "(" + "+".join(f'({faixa}="{l}")' for l in letras) + ")"


# grupos de códigos pelo turno EFETIVAMENTE trabalhado (M+ é manhã, Tm- é tarde…)
def codigos_por_efetivo(*bases):
    return [k for k in D.TURNOS if D.efetivo(k) in bases]


def expr_ausencias(lin, spans):
    """quantos dias de férias/licença/abono nos trechos de coluna da semana."""
    partes = []
    for a2, b2 in spans:
        faixa = f"${get_column_letter(a2)}{lin}:${get_column_letter(b2)}{lin}"
        partes += [f'COUNTIF({faixa},"{x}")' for x in sorted(D.AUSENCIAS)]
    return "+".join(partes)


def _outubro_ao_vivo(roster):
    """lê a aba OUT (e o 01/11 da NOV) do Sheet antigo agora — o que a Mari tem lá
    é o dado; só os códigos renomeados (E→CEP, P→CP, R→CRO) mudam de grafia."""
    import gsuite
    out = gsuite.ler(SHEET_ANTIGO, "OUT!A8:AM73")
    nov = gsuite.ler(SHEET_ANTIGO, "NOV!A8:I73")
    por_pessoa = {}
    for row in out:
        if not row or row[0] not in roster:
            continue
        for d in range(1, 32):
            idx = 8 + d - 1
            v = (row[idx] if idx < len(row) else "").strip()
            if v:
                por_pessoa.setdefault(row[0], {})[d] = RENOMEADOS.get(v, v)
    for row in nov:
        if row and row[0] in roster and len(row) > 8 and row[8].strip():
            por_pessoa.setdefault(row[0], {})[32] = RENOMEADOS.get(row[8].strip(), row[8].strip())
    print(f"outubro ao vivo: {sum(len(v) for v in por_pessoa.values())} células lidas do Sheet antigo")
    return por_pessoa


def carregar_dados():
    """DIAS[date][apelido] = (letra, origem) pro ano inteiro — SÓ fontes reais."""
    roster = [x[0] for x in D.ROSTER]
    DIAS, _ = senior_import.importar()
    grade, rel_grade = grade_import.importar(roster + [x[0] for x in D.FORA_DO_ROSTER])
    if MODO == "antiga":
        # dados como estavam: Senior manda, grade só onde não há Senior (jan/mai/jun)
        for data, pessoas in grade.items():
            if data.month in (1, 5, 6):
                for p_, cel in pessoas.items():
                    DIAS.setdefault(data, {}).setdefault(p_, cel)
        por_pessoa = _outubro_ao_vivo(roster)
        for data in [dt.date(2026, 10, d) for d in range(1, 32)] + [dt.date(2026, 11, 1)]:
            dia_k = 32 if data.month == 11 else data.day
            for apelido in roster:
                letra = por_pessoa.get(apelido, {}).get(dia_k)
                if letra:
                    DIAS.setdefault(data, {})[apelido] = (letra, "mari")
                else:
                    DIAS.get(data, {}).pop(apelido, None)
        return DIAS, rel_grade
    ESTRUTURAIS = {"C", "J", "CEP", "A"}       # o que só o arquivo Senior distingue
    # A GRADE DO GRUPO é a fonte de todo mês que tem grade (Marcos, 02/09/26: "o
    # Senior é uma coisa à parte que nem sempre condiz com a realidade"). Do
    # Senior ficam só os códigos estruturais de quem a grade lista como M/T/D e
    # as ausências/administrativo de quem ela não lista. Dia sem grade (01/01)
    # fica com o que o Senior tiver.
    for data, pessoas in grade.items():
        sen = DIAS.get(data, {})
        novo = {}
        for p, cel in pessoas.items():
            s_ = sen.get(p)
            if s_ and s_[0] in ESTRUTURAIS and cel[0] in ("M", "T", "D"):
                novo[p] = (s_[0], "senior")
            else:
                novo[p] = cel
        for p, s_ in sen.items():
            if p not in novo and (s_[0] in D.AUSENCIAS or s_[0] == "A"):
                novo[p] = s_
        DIAS[data] = novo
    # OUTUBRO: a versão VIVA da Mari (re-transcrita em 01/09/26) + as correções
    # da checagem dela e do Marcos. Substituição total da matriz: célula que ela
    # deixou vazia fica vazia (ela preferiu buraco a convocação).
    import mari_out_dados as MO
    import checagem_out_v3 as CK
    por_pessoa = {ap: dict(d) for ap, d in MO.MARI.items()}
    mexidos = {(ap, dia) for _i, ap, dia, _de, _para, _m in CK.EDITS}
    CK.aplicar(por_pessoa, verbose=True)
    for data in [dt.date(2026, 10, d) for d in range(1, 32)] + [dt.date(2026, 11, 1)]:
        dia_k = 32 if data.month == 11 else data.day
        for apelido in roster:
            letra = por_pessoa.get(apelido, {}).get(dia_k)
            if letra:
                origem = "checagem" if (apelido, dia_k) in mexidos else "mari"
                DIAS.setdefault(data, {})[apelido] = (letra, origem)
            else:
                DIAS.get(data, {}).pop(apelido, None)
    return DIAS, rel_grade


# ============================================================== abas
def altura(texto, largura, tam=9, minimo=15, extra=0):
    """altura de linha (pt) pra um texto com quebra numa largura de coluna.
    O Sheets não auto-ajusta linha com célula mesclada nem linha com altura
    fixa — então toda linha com texto longo ganha altura calculada."""
    import math
    por_linha = max(1.0, largura * {8: 1.35, 9: 1.2, 10: 1.05}.get(tam, 1.1))
    linhas = sum(max(1, math.ceil(len(p) / por_linha)) for p in str(texto).split("\n"))
    return max(minimo, (linhas + extra) * (tam * 1.45) + 5)


def altura_bloco(ws, r1, r2, texto, largura, tam=9):
    """distribui a altura de um texto mesclado em várias linhas (r1..r2)."""
    total = altura(texto, largura, tam)
    n = r2 - r1 + 1
    for rr in range(r1, r2 + 1):
        atual = ws.row_dimensions[rr].height or 15
        ws.row_dimensions[rr].height = max(atual, total / n)


VERTICAL = Alignment(horizontal="center", vertical="top", textRotation=90)


def _tip(cel, chave):
    """cola o tooltip do dicionário central como comentário (nota no Sheets)."""
    texto = D.TOOLTIPS.get(chave) or D.TOOLTIPS.get(str(chave).lower())
    if texto:
        cm = openpyxl.comments.Comment(texto, "colo ritmo")
        cm.width, cm.height = 320, 130
        cel.comment = cm
    return cel


def estilo_titulo(ws, texto, sub=""):
    c = ws.cell(row=1, column=1, value=texto)
    c.font = Font(name=DISPLAY, bold=True, size=15, color=INK)
    if sub:
        s = ws.cell(row=2, column=1, value=sub)
        s.font = Font(name=F, size=9, italic=True, color=INK3)
    ws.sheet_view.showGridLines = False


def creme(ws, ate_linha, ate_coluna):
    """pinta o creme por baixo de tudo — '#FFFAF3 · fundo de TUDO' nos tokens.

    Roda ANTES dos preenchimentos específicos não daria certo (a ordem de escrita
    é outra), então só pinta célula que ficou sem fill próprio.
    """
    vazio = ("00000000", None)
    for r in range(1, ate_linha + 1):
        for c in range(1, ate_coluna + 1):
            cel = ws.cell(row=r, column=c)
            if cel.fill is None or cel.fill.fgColor.rgb in vazio:
                cel.fill = PatternFill("solid", fgColor=CREME)


def aba_leiame(wb, rel_grade):
    ws = wb.create_sheet("LEIA-ME")
    ws.sheet_properties.tabColor = PINK
    estilo_titulo(ws, "Escala UTI HCB · 2026 — V3" if MODO == "v3" else
                  "Escala UTI HCB · 2026 — unificada (estrutura da V3, dados desta planilha)")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 17
    r = 3

    def secao(titulo):
        nonlocal r
        c = ws.cell(row=r, column=1, value=titulo)
        c.font = Font(name=DISPLAY, bold=True, size=11, color=LAVI)
        r += 1

    def linha(a, b="", cor=INK2, negrito=False):
        nonlocal r
        ca = ws.cell(row=r, column=1, value=a)
        ca.font = Font(name=F, size=10, bold=negrito, color=INK)
        ca.alignment = Alignment(vertical="top")
        if b:
            cel = ws.cell(row=r, column=2, value=b)
            cel.font = Font(name=F, size=10, color=cor)
            cel.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
            ws.row_dimensions[r].height = altura(b, 132, 10)
        r += 1

    secao("O que é")
    linha("", "O ano inteiro num arquivo: 12 abas mensais com a escala real, o painel "
              "comparativo do ano, a tradução pros códigos do Senior e o validador das "
              "regras que são lei. Tudo o que é contagem é fórmula — digitou o código na "
              "aba do mês, a cobertura, a carga horária, o saldo e os alertas se refazem sozinhos.")
    r += 1
    secao("Legenda dos códigos")
    ws.cell(row=r, column=1, value="código").font = Font(name=F, bold=True, size=9, color="FFFFFF")
    for i, t in enumerate(("turno", "horário", "horas", "Senior")):
        ws.cell(row=r, column=2 + i, value=t).font = Font(name=F, bold=True, size=9, color="FFFFFF")
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LAVI)
    r += 1
    for letra, (rot, hora, horas, cod, *_) in D.TURNOS.items():
        cel = ws.cell(row=r, column=1, value=letra)
        base = D.base_visual(letra)
        cel.font = Font(name=F, bold=True, size=10, italic=letra in D.BHN,
                        strike=letra in D.BHN,
                        color=(CORALI if letra in D.BHP else
                               (INK3 if letra in D.BHN else
                                ("FFFFFF" if base in BRANCO else INK))))
        cel.fill = PatternFill("solid", fgColor=(FILL_BHP if letra in D.BHP else
                                                 (FILL_BHN if letra in D.BHN else
                                                  FILLS.get(base, CREME))))
        cel.alignment = Alignment(horizontal="center")
        for i, v in enumerate((rot, hora, horas if horas else "—", cod or "—")):
            ws.cell(row=r, column=2 + i, value=v).font = Font(name=F, size=10, color=INK2)
        r += 1
    linha("Antigos", "Até abril de 2026 os códigos eram outros: "
          + " · ".join(f"{k} = {v}" for k, v in D.CODIGOS_ANTIGOS.items())
          + ". A mudança de abril passou a usar 2/40/41.")
    linha("Banco", "Código com + é BHP: a pessoa trabalha a MAIS e a hora fica no banco — o "
          "Senior não recebe esse plantão. Código com − é BHN: dispensa paga pelo banco — a "
          "pessoa não vem, mas o Senior recebe o plantão normal. Dm+ é um dia inteiro em que "
          "só a manhã é BHP; Tm- é uma tarde trabalhada com a manhã dispensada.", negrito=True)
    r += 1

    secao("As abas")
    for nome, desc in (
        ("CADASTRO", "Quem é quem: carga horária, restrições duras, sexta-noite e fds extra. "
                     "É a fonte dos nomes de todas as abas mensais."),
        ("CONFIG", "As regras como dado: cobertura mínima, cota de fds em mês com férias, "
                   "tabela de códigos e as regras duras com o artigo da CLT. Mudou aqui, "
                   "mudou na planilha toda."),
        ("JAN a DEZ", "A escala: matriz médico × dia, com filtro no cabeçalho pra ordenar "
                      "(veja “Ordenar” abaixo). A semana fecha NO DOMINGO, então todo "
                      "mês aparece com as semanas completas: as primeiras colunas trazem o fim "
                      "do mês anterior e as últimas trazem os dias do mês seguinte até o domingo "
                      "que fecha a última semana. Esses dias vizinhos são fórmula da aba deles — "
                      "pra editar, use a aba do mês dono do dia. À direita, por pessoa, PRIMEIRO "
                      "as semanas: Sem 1..6 = horas de cada semana civil (o intervalo de datas "
                      "está em cima) e, ao lado de cada uma, BH = horas − alvo semanal da pessoa "
                      "(+ é BHP e fica areia, − é BHN/faltou e fica coral; zero é semana fechada). "
                      "Depois: carga do mês, fds, sexta-noite, feriado, meta, saldo e os alertas "
                      "de jornada. No rodapé, a lotação de cada turno e quanto falta pro mínimo."),
        ("Ordenar", "Cada aba mensal tem um FILTRO no cabeçalho (ícone de funil): clique no "
                    "funil de “Médico” e ordene de A a Z, ou no de “Grupo” "
                    "para ver coordenação → rotina → staff. A coluna “Nº” volta à "
                    "ordem original. A planilha foi preparada para isso: toda fórmula acha "
                    "cada pessoa PELO NOME, em qualquer ordem — ordenar pelo funil não "
                    "quebra nada. Só não use “Classificar intervalo” numa seleção "
                    "parcial, que aí as linhas se separam das colunas."),
        ("OUT · DIA A DIA", "O mês em formato calendário: por dia e por turno, quem "
                             "está onde, e a cobertura. Abre na segunda-feira da semana do "
                             "dia 1º e fecha no domingo da última semana — os dias dos meses "
                             "vizinhos aparecem acinzentados, só para leitura. É fórmula da "
                             "aba do mês — trocou um plantão no dropdown, o dia a dia se "
                             "refaz. Bom pra bater o olho e perceber o que está esquisito."),
        ("PAINEL ANO", "O comparativo do ano: saldo, fds, sexta-noite e feriado de cada pessoa "
                       "mês a mês, com acumulado. Onde existe a contagem manual antiga, ela "
                       "aparece ao lado pra conferência."),
        ("SENIOR", "Escolha o mês em B1 e a matriz inteira sai traduzida nos códigos do RH, "
                   "pronta pra lançar."),
        ("VALIDADOR", "O que fere regra dura, separado em ESTRUTURAL (é a forma do contrato, "
                      "decisão de política) e PONTUAL (erro daquele mês)."),
        ("CHECAGEM OUT", "A checagem de outubro feita por Marcos e Mari: as regras que ela "
                         "revelou, cada célula que mudou (estava → ficou, por quê) e o que foi "
                         "verificado sem mudar."),
    ):
        linha(nome, desc, negrito=True)
    r += 1

    secao("De onde vem cada mês — leia antes de comparar")
    for mes in range(1, 13):
        origem, nota = D.PROCEDENCIA[mes]
        cor = {"senior": SAGES, "grade": SANDS, "montado": LAVS, "vazio": LINE}[origem]
        cel = ws.cell(row=r, column=1, value=MESES_PT[mes - 1])
        cel.font = Font(name=F, bold=True, size=10, color=INK)
        cel.fill = PatternFill("solid", fgColor=cor)
        cel.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=r, column=2, value=nota)
        c2.font = Font(name=F, size=9, color=INK2)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.row_dimensions[r].height = altura(nota, 132, 9)
        r += 1
    r += 1
    linha("atenção", D.AVISO_GRADE, cor=CORALI)
    r += 1
    secao("O que ainda falta")
    for falta in (
        "Janeiro começa no dia 2: a grade recebida não tem o dia 1º.",
        "Toda a escala de janeiro a setembro vem das GRADES do grupo. O arquivo Senior só "
        "empresta os códigos 47/78/6/11 e as férias — janeiro, maio e junho não têm Senior, "
        "então a chefia desses meses aparece como manhã de 6h.",
        "Novembro e dezembro estão em branco, prontos pra montar (01/11 já tem o que a Mari lançou).",
        "Os dias do CRO da LuAlice em outubro (02 e 09) e o domingo 18/10 à noite (ficou com 6/7 "
        "depois de desfazer as 24h emendadas do Moabe) estão marcados pra decisão da Mari na aba "
        "CHECAGEM OUT.",
        "Dois códigos órfãos no histórico: 23 (Isabella 23/08) e 3 (Laura 02/08) — não "
        "constam em nenhuma legenda; ficaram de fora.",
        "A coluna 18h⚠ enxerga a virada de mês pelas colunas de véspera: a noite do fim "
        "do mês anterior emendando a manhã do dia 1º conta no mês NOVO (cada virada tem "
        "um dono só, nunca é contada duas vezes). A virada dez/26→jan/27 fica com a aba "
        "VALIDADOR, que olha o ano inteiro.",
        "O seletor de mês da aba SENIOR usa INDIRECT, que só resolve com o arquivo aberto "
        "no Excel ou no Google Sheets — em pré-visualização de e-mail ou no Drive sem abrir, "
        "a matriz aparece vazia. Abrir e escolher o mês em B1 resolve.",
    ):
        linha("·", falta)
    creme(ws, r + 2, 8)
    return ws


def aba_cadastro(wb):
    ws = wb.create_sheet("CADASTRO")
    ws.sheet_properties.tabColor = AQUA
    estilo_titulo(ws, "Cadastro · Quem é quem",
                  f"{len(D.ROSTER)} pessoas ativas")
    cabecalhos = ["Médico", "Nome completo", "CH", "Grupo", "Restrições duras",
                  "Sexta-noite", "FDS extra", "Observações", "Alvo/sem"]
    larguras = [14, 22, 5, 15, 58, 16, 14, 52, 8]
    chaves_cad = {"CH": "CH", "Grupo": "grupo", "Sexta-noite": "sexta-noite ficha",
                  "FDS extra": "fds extra ficha", "Médico": "médico", "Alvo/sem": "alvo-sem"}
    for i, (h, w) in enumerate(zip(cabecalhos, larguras), start=1):
        cel = ws.cell(row=3, column=i, value=h)
        cel.font = Font(name=F, bold=True, size=9, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor=LAVI)
        cel.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = w
        if h in chaves_cad:
            _tip(cel, chaves_cad[h])
    grupos = {"chefia": "FAEAF2", "rotina": "FAEAF2", "30h": AQUAS,
              "36h": SANDS, "24h": CREME, "administrativo": LINE}
    r = 4
    for apelido, nome, ch, grupo, restr, sn, fe, obs in D.ROSTER:
        vals = [apelido, nome, ch, grupo, restr, sn, fe, obs, D.ALVO_SEMANAL.get(apelido, ch)]
        for i, v in enumerate(vals, start=1):
            cel = ws.cell(row=r, column=i, value=v)
            cel.font = Font(name=F, size=9, color=INK, bold=(i in (1, 9)))
            cel.alignment = Alignment(wrap_text=(i in (5, 8)), vertical="top")
            cel.border = BOX
            if i == 4:
                cel.fill = PatternFill("solid", fgColor=grupos.get(grupo, CREME))
        if (r % 2) == 1:                      # banda quente alternada
            for i in range(1, 10):
                cel = ws.cell(row=r, column=i)
                if cel.fill is None or cel.fill.fgColor.rgb in ("00000000", None):
                    cel.fill = PatternFill("solid", fgColor=CREME2)
        ws.row_dimensions[r].height = max(altura(restr, 58), altura(obs, 52), 18)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Fora do roster").font = Font(name=F, bold=True, size=10, color=CORALI)
    r += 1
    for apelido, motivo in D.FORA_DO_ROSTER:
        ws.cell(row=r, column=1, value=apelido).font = Font(name=F, size=9, bold=True, color=INK3)
        ws.cell(row=r, column=2, value=motivo).font = Font(name=F, size=9, color=INK3)
        r += 1
    creme(ws, r + 2, 9)
    ws.freeze_panes = "C4"
    return ws


def aba_config(wb):
    ws = wb.create_sheet("CONFIG")
    ws.sheet_properties.tabColor = SAGE
    estilo_titulo(ws, "Config · As regras como dado",
                  "Mudou aqui, muda na planilha toda")
    for col, w in zip("ABCDEFG", (18, 13, 13, 13, 13, 44, 44)):
        ws.column_dimensions[col].width = w

    def cab(r, textos):
        for i, t in enumerate(textos, start=1):
            cel = ws.cell(row=r, column=i, value=t)
            cel.font = Font(name=F, bold=True, size=9, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor=LAVI)

    # cobertura mínima — DOIS blocos, porque a regra mudou em outubro/26.
    # Linhas 4-6 = regra em vigor · linhas 9-11 = regra antiga (jan–set).
    # As abas mensais apontam para o bloco da vigência do próprio mês.
    ws.cell(row=3, column=1, value="Cobertura mínima por turno").font = Font(
        name=DISPLAY, bold=True, size=11, color=LAVI)
    cab(3, ["Em vigor · de out/26", "Manhã", "Tarde", "Noite"])
    for i, tipo in enumerate(("útil", "sábado", "domingo")):
        m, t, n = D.MINIMOS[tipo]
        ws.cell(row=4 + i, column=1, value=tipo).font = Font(name=F, size=10, bold=True, color=INK)
        for j, v in enumerate((m, t, n)):
            cel = ws.cell(row=4 + i, column=2 + j, value=v)
            cel.font = Font(name=F, size=11, bold=True, color=LAVI)
            cel.alignment = Alignment(horizontal="center")
            cel.border = BOX
    nota = ws.cell(row=4, column=6, value=D.NOTA_MINIMOS)
    nota.font = Font(name=F, size=9, color=INK2)
    nota.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=4, start_column=6, end_row=7, end_column=7)
    altura_bloco(ws, 4, 7, D.NOTA_MINIMOS, 88)
    cab(8, ["Até set/26 · histórico", "Manhã", "Tarde", "Noite"])
    for i, tipo in enumerate(("útil", "sábado", "domingo")):
        m, t, n = D.MINIMOS_ANTIGOS[tipo]
        ws.cell(row=9 + i, column=1, value=tipo).font = Font(name=F, size=10, color=INK2)
        for j, v in enumerate((m, t, n)):
            cel = ws.cell(row=9 + i, column=2 + j, value=v)
            cel.font = Font(name=F, size=10, color=INK2)
            cel.alignment = Alignment(horizontal="center")
            cel.border = BOX
    nota2 = ws.cell(row=9, column=6, value="Os meses de janeiro a setembro são medidos "
                    "por ESTA regra. Cobrar deles o mínimo novo seria exigir do passado "
                    "uma regra que ainda não existia.")
    nota2.font = Font(name=F, size=9, italic=True, color=INK3)
    nota2.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=9, start_column=6, end_row=11, end_column=7)
    altura_bloco(ws, 9, 11, nota2.value, 88)

    # tabela de códigos — linhas 9..20, referenciada pela aba SENIOR
    ws.cell(row=13, column=1, value="Tabela de códigos").font = Font(
        name=DISPLAY, bold=True, size=11, color=LAVI)
    cab(14, ["Letra", "Senior", "Horas", "Conta manhã", "Conta tarde", "Conta noite", "Turno"])
    for cc in range(4, 7):
        _tip(ws.cell(row=14, column=cc), "conta-flags")
    r = 15
    for letra, (rot, hora, horas, cod, cm, ct, cn) in D.TURNOS.items():
        ws.cell(row=r, column=1, value=letra).font = Font(name=F, bold=True, size=10, color=INK)
        ws.cell(row=r, column=2, value=cod or "")
        ws.cell(row=r, column=3, value=horas)
        ws.cell(row=r, column=4, value=cm)
        ws.cell(row=r, column=5, value=ct)
        ws.cell(row=r, column=6, value=cn)
        ws.cell(row=r, column=7, value=f"{rot} · {hora}")
        for c in range(1, 8):
            ws.cell(row=r, column=c).font = Font(
                name=F, size=9, color=INK, bold=(c == 1))
            ws.cell(row=r, column=c).border = BOX
        r += 1
    fim_cod = r - 1

    # cota de fds em férias
    r += 1
    ws.cell(row=r, column=1, value="Cota de FDS em mês com férias").font = Font(
        name=DISPLAY, bold=True, size=11, color=LAVI)
    r += 1
    cab(r, ["CH semanal", "Férias 2 sem", "Férias 1 sem", "Sem férias"])
    r += 1
    r_cota = r                      # as abas mensais apontam para cá
    fl = ws.cell(row=r - 2, column=6, value=(
        "Em mês de 5 fins de semana a demanda de fds passa da soma das cotas do "
        "grupo e zerar o excesso é aritmeticamente impossível. Outubro/26: demanda "
        "1704h contra 1494h de cota somada — 210h que não têm de onde sair. O alvo "
        "de cada pessoa passa a ser a cota vezes o fator abaixo, para que o excesso "
        "inevitável seja dividido em proporção e não despejado em quem tem menos veto."))
    fl.font = Font(name=F, size=9, italic=True, color=INK2)
    fl.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r - 2, start_column=6, end_row=r + 2, end_column=7)
    altura_bloco(ws, r - 2, r + 2, fl.value, 88)
    cfat = ws.cell(row=r - 1, column=5, value="Fator do mês")
    cfat.font = Font(name=F, size=9, bold=True, color=INK)
    _tip(cfat, "fator")
    cf_fator = ws.cell(row=r, column=5, value=1.141)
    cf_fator.font = Font(name=F, size=11, bold=True, color=LAVI)
    cf_fator.number_format = "0.000"
    cf_fator.alignment = Alignment(horizontal="center")
    cf_fator.border = BOX
    for ch, duas, uma, sem in D.COTA_FDS_FERIAS:
        for j, v in enumerate((ch, duas, uma, sem)):
            cel = ws.cell(row=r, column=1 + j, value=v)
            cel.font = Font(name=F, size=10, color=INK, bold=(j == 0))
            cel.alignment = Alignment(horizontal="center")
            cel.border = BOX
        r += 1

    # alvo semanal e banco de horas (regra revelada pela checagem de 28/08)
    r += 1
    ws.cell(row=r, column=1, value="Alvo semanal e banco de horas").font = Font(
        name=DISPLAY, bold=True, size=11, color=LAVI)
    na = ws.cell(row=r, column=6, value=D.NOTA_ALVO)
    na.font = Font(name=F, size=9, color=INK2)
    na.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=6, end_row=r + 3, end_column=7)
    altura_bloco(ws, r, r + 3, D.NOTA_ALVO, 88)
    r += 1
    cab(r, ["Parâmetro", "Valor"])
    r += 1
    r_desc = r                      # as abas mensais apontam para cá (BH k)
    ws.cell(row=r, column=1, value="Desconto por dia de férias/licença/abono na semana (h)").font = Font(
        name=F, size=9, bold=True, color=INK)
    cd = ws.cell(row=r, column=2, value=D.DESCONTO_AUSENCIA_SEMANA)
    cd.font = Font(name=F, size=11, bold=True, color=LAVI)
    cd.alignment = Alignment(horizontal="center")
    cd.border = BOX
    r += 1
    ws.cell(row=r, column=1, value="Alvo por pessoa: CADASTRO, coluna Alvo/sem").font = Font(
        name=F, size=9, italic=True, color=INK3)
    r += 1

    # feriados do ano — eram dado escondido no Python; agora vivem aqui
    r += 1
    ws.cell(row=r, column=1, value="Feriados de 2026").font = Font(
        name=DISPLAY, bold=True, size=11, color=LAVI)
    nf = ws.cell(row=r, column=5, value=D.NOTA_FERIADOS)
    nf.font = Font(name=F, size=9, color=INK2)
    nf.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=5, end_row=r + 5, end_column=7)
    altura_bloco(ws, r, r + 5, D.NOTA_FERIADOS, 101)
    r += 1
    cab(r, ["Data", "Sigla", "Feriado"])
    r += 1
    for (mes, dia), (nome_f, sigla) in sorted(D.FERIADOS_2026.items()):
        data = dt.date(2026, mes, dia)
        cel = ws.cell(row=r, column=1, value=f"{dia:02d}/{mes:02d} ({WD_PT[data.weekday()]})")
        cel.font = Font(name=F, size=9, bold=True, color=INK)
        sg = ws.cell(row=r, column=2, value=sigla)
        sg.font = Font(name=F, size=9, bold=True, color="FFFFFF")
        sg.fill = PatternFill("solid", fgColor=CORALI)
        sg.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=nome_f).font = Font(name=F, size=9, color=INK2)
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = BOX
        r += 1

    # regras duras
    r += 1
    ws.cell(row=r, column=1, value="Regras duras").font = Font(name=DISPLAY, bold=True, size=11, color=LAVI)
    r += 1
    cab(r, ["Regra", "O que é", "Base", "Tratamento", "Nota"])
    ws.cell(row=r, column=2).value = "o que é"
    r += 1
    cores = {"ALERTA": CORALS, "REGISTRO": OLIVES, "CÁLCULO": AQUAS, "PREFERÊNCIA": CREME}
    for nome, oque, base, trat, nota_txt in D.REGRAS_DURAS:
        ws.cell(row=r, column=1, value=nome).font = Font(name=F, size=9, bold=True, color=INK)
        for i, v in enumerate((oque, base, trat), start=2):
            cx = ws.cell(row=r, column=i, value=v)
            cx.font = Font(name=F, size=9, color=INK2)
            cx.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor=cores.get(trat, CREME))
        cel = ws.cell(row=r, column=5, value=nota_txt)
        cel.font = Font(name=F, size=9, color=INK2)
        cel.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        ws.row_dimensions[r].height = max(altura(nota_txt, 101), altura(oque, 13), altura(trat, 13))
        r += 1
    creme(ws, r + 2, 7)
    return ws, fim_cod, r_cota, r_desc


def aba_mes(wb, mes, DIAS, fim_cod, pessoas, r_cota, r_desc):
    nome = MESES_PT[mes - 1]
    ws = wb.create_sheet(nome)
    origem, nota = D.PROCEDENCIA[mes]
    if MODO == "antiga":
        nota = {10: "ESCALA DA MARI — os dados desta planilha, lidos ao vivo na hora de gerar a "
                    "estrutura nova (02/09/26). Nenhuma correção da checagem foi aplicada aqui; "
                    "elas estão na planilha V3", }.get(mes, "Fontes de antes: códigos Senior; grade "
                    "do grupo em janeiro, maio e junho" if mes <= 9 else nota)
    ws.sheet_properties.tabColor = {"senior": SAGE, "grade": SAND,
                                   "montado": LAV, "vazio": LINE}[origem]
    ws.sheet_view.showGridLines = False
    ndias = (dt.date(2026, mes % 12 + 1, 1) - dt.timedelta(days=1)).day if mes < 12 else 31

    t = ws.cell(row=R_TIT, column=1, value=f"{MESES_LONGO[mes-1].capitalize()} · 2026")
    t.font = Font(name=DISPLAY, bold=True, size=14, color=INK)
    t.comment = openpyxl.comments.Comment(f"De onde vem este mês: {nota}", "colo ritmo")
    # legenda compacta sempre à vista — pra ninguém precisar decorar as siglas
    sub = ws.cell(row=R_TIT, column=C_D1 + 6, value=(
        "Legenda: M manhã 7–13 · T tarde 13–19 · D dia 7–19 · N noite 19–7 · "
        "NT noitinha · C chefia 10h · J Janaina 8–13 · CEP · CP · CRO · "
        "+ = BHP (a mais, no banco) · − = BHN (dispensa) · FE férias · LM licença · AB abono"))
    sub.font = Font(name=F, size=8, italic=True, color=INK3)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 5
    # vésperas: o fim do mês anterior que fecha a 1ª semana civil (seg–dom).
    # Valores vêm POR FÓRMULA da aba anterior — trocou lá, atualiza aqui.
    # Máscaras zeradas: véspera não conta em CH/FDS/saldo deste mês.
    offset = dt.date(2026, mes, 1).weekday()       # quantas vésperas o mês pede
    aba_prev = MESES_PT[mes - 2] if mes > 1 else None
    # virada: a semana fecha NO DOMINGO — os dias do mês seguinte até esse
    # domingo aparecem depois do dia final, por fórmula da aba seguinte
    n_virada = (7 - ((offset + ndias) % 7)) % 7
    aba_prox = MESES_PT[mes % 12] if mes < 12 else None
    prox1 = dt.date(2026 + (1 if mes == 12 else 0), mes % 12 + 1, 1)
    dias_virada = [prox1 + dt.timedelta(days=j) for j in range(n_virada)]
    for j in range(N_POS):
        col = C_D32 + j
        for rr in (R_FDS, R_SEX, R_FER):
            ws.cell(row=rr, column=col, value=0)   # cota é do mês seguinte
        if j >= n_virada:
            ws.column_dimensions[get_column_letter(col)].width = 2.5
            for rr in (R_DIA, R_DOW):
                ws.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=LINE)
            continue
        d_prox = dias_virada[j]
        ws.column_dimensions[get_column_letter(col)].width = 5.3
        cd = ws.cell(row=R_DIA, column=col, value=d_prox.strftime("%d/%m"))
        cd.font = Font(name=F, bold=True, size=7, color=INK2)
        cd.alignment = Alignment(horizontal="center")
        cd.fill = PatternFill("solid", fgColor=LAVS if d_prox.weekday() >= 5 else CREME2)
        cw = ws.cell(row=R_DOW, column=col, value=WD_PT[d_prox.weekday()])
        cw.font = Font(name=F, size=8, color=CORALI if d_prox.weekday() >= 5 else INK3)
        cw.alignment = Alignment(horizontal="center")
        _tip(cd, "dia-seguinte")
    for j in range(N_PRE):
        col = C_PRE1 + j
        usada = j >= N_PRE - offset
        for rr in (R_FDS, R_SEX, R_FER):
            ws.cell(row=rr, column=col, value=0)
        if not usada:
            ws.column_dimensions[get_column_letter(col)].width = 2.5
            for rr in (R_DIA, R_DOW):
                ws.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=LINE)
            continue
        ws.column_dimensions[get_column_letter(col)].width = 5.3
        d_prev = dt.date(2026, mes, 1) - dt.timedelta(days=N_PRE - j)
        cd = ws.cell(row=R_DIA, column=col, value=d_prev.strftime("%d/%m"))
        cd.font = Font(name=F, bold=True, size=7, color=INK2)
        cd.alignment = Alignment(horizontal="center")
        cd.fill = PatternFill("solid", fgColor=LAVS if d_prev.weekday() >= 5 else CREME2)
        cw = ws.cell(row=R_DOW, column=col, value=WD_PT[d_prev.weekday()])
        cw.font = Font(name=F, size=8, color=CORALI if d_prev.weekday() >= 5 else INK3)
        cw.alignment = Alignment(horizontal="center")
        _tip(cd, "dia-anterior")
    # cabeçalho de dias + máscaras
    for i in range(31):
        col = C_D1 + i
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 5.3
        if i >= ndias:
            for r in (R_DIA, R_DOW):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LINE)
            continue
        data = dt.date(2026, mes, i + 1)
        fds = 1 if data.weekday() >= 5 else 0
        info_fer = D.FERIADOS_2026.get((mes, i + 1))
        feriado = 1 if info_fer else 0
        cd = ws.cell(row=R_DIA, column=col, value=i + 1)
        cd.font = Font(name=F, bold=True, size=9, color=INK)
        cd.alignment = Alignment(horizontal="center")
        # no feriado a coluna mostra a sigla do feriado em vez do dia da semana:
        # é a informação que muda a decisão, e o dia da semana se deduz do número
        rotulo = info_fer[1] if feriado else WD_PT[data.weekday()]
        cw = ws.cell(row=R_DOW, column=col, value=rotulo)
        cw.font = Font(name=F, size=8, bold=bool(feriado),
                       color="FFFFFF" if feriado else (CORALI if fds else INK3))
        cw.alignment = Alignment(horizontal="center")
        if feriado:
            cd.fill = PatternFill("solid", fgColor=CORAL)
            cd.font = Font(name=F, bold=True, size=9, color="FFFFFF")
            cw.fill = PatternFill("solid", fgColor=CORALI)
            cd.comment = openpyxl.comments.Comment(
                f"{info_fer[0]} — feriado.\n\nEscala como {WD_PT[data.weekday()]}: "
                f"a lotação exigida NÃO cai.", "colo ritmo")
        elif fds:
            cd.fill = PatternFill("solid", fgColor=LAVS)
        ws.cell(row=R_FDS, column=col, value=fds)
        ws.cell(row=R_SEX, column=col, value=1 if data.weekday() == 4 else 0)
        ws.cell(row=R_FER, column=col, value=feriado)
    for r in (R_FDS, R_SEX, R_FER):
        ws.row_dimensions[r].hidden = True
        ws.cell(row=r, column=1, value={R_FDS: "máscara fds", R_SEX: "máscara sexta",
                                        R_FER: "máscara feriado"}[r])

    # cabeçalho da matriz
    for col, txt in ((C_MED, "Médico"), (C_CH, "CH")):
        c = ws.cell(row=R_HDR, column=col, value=txt)
        c.font = Font(name=F, bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=LAVI)
        _tip(c, txt)
    jans = janelas_semana_civil(mes, ndias)
    for i, txt in enumerate(COLS_TOT):
        col = C_TOT + i
        c = _tip(ws.cell(row=R_HDR, column=col, value=txt),
                 "sem-n" if txt.startswith("Sem ") else
                 ("bh-n" if txt.startswith("BH ") else
                  ("grupo-filtro" if txt == "Grupo" else
                   ("ordem-original" if txt == "Nº" else txt))))
        c.font = Font(name=F, bold=True, size=8, color="FFFFFF")
        eh_sem = txt.startswith("Sem ") or txt.startswith("BH ")
        c.fill = PatternFill("solid", fgColor=(CORALI if txt in COLS_ALERTA else
                                                ("453A73" if eh_sem else LAVI)))
        c.alignment = VERTICAL          # todo o cabeçalho de totais na vertical (Marcos, 02/09)
        ws.column_dimensions[get_column_letter(col)].width = \
            (13 if txt == "Grupo" else (4.6 if eh_sem or txt == "Nº" else 5.4))
    ws.row_dimensions[R_HDR].height = 66
    # o intervalo de datas de cada semana, em cima do par Sem/BH
    d1 = dt.date(2026, mes, 1)
    for k, spans in enumerate(jans, start=1):
        ini_sem = d1 + dt.timedelta(days=(spans[0][0] - C_D1))
        fim_sem = ini_sem + dt.timedelta(days=6)
        c1 = col_tot(f"Sem {k}")
        ws.merge_cells(start_row=R_DIA, start_column=c1, end_row=R_DIA, end_column=c1 + 1)
        cd = ws.cell(row=R_DIA, column=c1, value=f"{ini_sem:%d/%m}–{fim_sem:%d/%m}")
        cd.font = Font(name=F, size=7, color=INK3)
        cd.alignment = Alignment(horizontal="center")
        cs = ws.cell(row=R_DOW, column=c1, value="semana civil")
        cs.font = Font(name=F, size=7, italic=True, color=INK3)
        ws.merge_cells(start_row=R_DOW, start_column=c1, end_row=R_DOW, end_column=c1 + 1)
        cs.alignment = Alignment(horizontal="center")

    semanas = round(ndias / 7, 2)
    for k, (apelido, ch) in enumerate(pessoas):
        lin = R_P0 + k
        cm = ws.cell(row=lin, column=C_MED, value=apelido)
        cm.font = Font(name=F, size=9, bold=True, color=INK)
        cc = ws.cell(row=lin, column=C_CH, value=ch)
        cc.font = Font(name=F, size=8, color=INK3)
        cc.alignment = Alignment(horizontal="center")
        for j in range(N_PRE):
            col = C_PRE1 + j
            c = ws.cell(row=lin, column=col)
            if j < N_PRE - offset or not aba_prev:
                c.fill = PatternFill("solid", fgColor=LINE)
                continue
            d_prev = dt.date(2026, mes, 1) - dt.timedelta(days=N_PRE - j)
            # busca PELO NOME (não pela posição): a aba pode ser ordenada pelo
            # filtro — em qualquer ordem, de cá ou de lá, a fórmula acha a pessoa
            c.value = _ref_vizinho(aba_prev, d_prev.day, lin)
            c.font = Font(name=F, size=9, color=INK2)
            c.alignment = Alignment(horizontal="center")
            c.border = BOX
            c.fill = PatternFill("solid",
                                 fgColor=LAVS if d_prev.weekday() >= 5 else CREME2)
        for i in range(ndias):
            col = C_D1 + i
            data = dt.date(2026, mes, i + 1)
            cel = DIAS.get(data, {}).get(apelido)
            c = ws.cell(row=lin, column=col)
            c.alignment = Alignment(horizontal="center")
            c.border = BOX
            if cel:
                letra = cel[0]
                c.value = letra
                base = D.base_visual(letra)
                if letra in D.BHN:          # dispensa: cinza, itálico, riscado
                    c.font = Font(name=F, size=8, italic=True, strike=True, color=INK3)
                    c.fill = PatternFill("solid", fgColor=FILL_BHN)
                elif letra in D.BHP:        # a mais: coral claro, negrito
                    c.font = Font(name=F, size=8, bold=True, color=CORALI)
                    c.fill = PatternFill("solid", fgColor=FILL_BHP)
                else:
                    c.font = Font(name=F, size=9, bold=True,
                                  color="FFFFFF" if base in BRANCO else INK)
                    if base in FILLS:
                        c.fill = PatternFill("solid", fgColor=FILLS[base])
            else:
                c.font = Font(name=F, size=9, color=INK)
        for i in range(ndias, 31):
            ws.cell(row=lin, column=C_D1 + i).fill = PatternFill("solid", fgColor=LINE)
        # virada: fórmulas vivas da aba do mês seguinte (editar é LÁ — aqui é
        # conferência da semana que fecha no domingo), busca pelo nome
        for j in range(N_POS):
            col = C_D32 + j
            c = ws.cell(row=lin, column=col)
            if j >= n_virada or not aba_prox:
                c.fill = PatternFill("solid", fgColor=LINE)
                continue
            d_prox = dias_virada[j]
            c.value = _ref_vizinho(aba_prox, d_prox.day, lin)
            c.font = Font(name=F, size=9, color=INK2)
            c.alignment = Alignment(horizontal="center")
            c.border = BOX
            c.fill = PatternFill("solid",
                                 fgColor=LAVS if d_prox.weekday() >= 5 else CREME2)

        eh = expr_horas(lin)
        L = lambda nome: get_column_letter(col_tot(nome))
        LC = get_column_letter(C_CH)
        noites = codigos_por_efetivo("N")
        manhas = codigos_por_efetivo("M", "D", "C", "CEP")
        tardes = codigos_por_efetivo("T")
        r1 = f"${L_PRE1}{lin}:${L_DN_1}{lin}"       # dia D
        r2 = f"${L_PRE2}{lin}:${L_DN}{lin}"         # dia D+1
        formulas = {
            "CH mês":  f"=SUMPRODUCT({eh})",
            "FDS":     f"=SUMPRODUCT(${L_D1}${R_FDS}:${L_DN}${R_FDS},{eh})",
            "SxN":     f"=SUMPRODUCT(${L_D1}${R_SEX}:${L_DN}${R_SEX},"
                       f"{expr_conta(f'${L_D1}{lin}:${L_DN}{lin}', noites)})",
            "Feriado": f"=SUMPRODUCT(${L_D1}${R_FER}:${L_DN}${R_FER},{eh})",
            "Meta":    f"=ROUND({LC}{lin}*{semanas},0)",
            "Saldo":   f"={L('CH mês')}{lin}-{L('Meta')}{lin}",
            # começa nas vésperas: a noite do fim do mês anterior emendando a
            # manhã do dia 1º é vista aqui (a do dia 31→1º seguinte é vista na
            # aba do mês seguinte — cada virada tem um dono só)
            "18h⚠":    f"=SUMPRODUCT({expr_conta(r1, noites)}*{expr_conta(r2, manhas)})",
            "N→T":     f"=SUMPRODUCT({expr_conta(r1, noites)}*{expr_conta(r2, tardes)})",
            # cota de fds: base pela CH, reduzida conforme as semanas de férias no mês.
            # A regra vive em CONFIG; aqui só se aponta pra lá. CH fora da tabela
            # (40h) fica em branco de propósito: o doc diz "40h segue caso a caso".
            "Cota FDS": f'=IFERROR(INDEX(CONFIG!$B${r_cota}:$D${r_cota+2},'
                        f'MATCH(${LC}{lin},CONFIG!$A${r_cota}:$A${r_cota+2},0),'
                        f'IF(COUNTIF(${L_D1}{lin}:${L_DN}{lin},"FE")>=10,1,'
                        f'IF(COUNTIF(${L_D1}{lin}:${L_DN}{lin},"FE")>=1,2,3))),"")',
            # alerta contra o ALVO proporcional (cota × fator do mês), não contra a
            # cota crua: em mês de 5 fds o excesso é inevitável e o que importa é
            # se a pessoa carrega mais do que a fatia dela
            "FDS⚠":    f'=IF({L("Cota FDS")}{lin}="","",'
                       f'MAX(0,{L("FDS")}{lin}-{L("Cota FDS")}{lin}*CONFIG!$E${r_cota}))',
            "Sem⚠":    "=MAX(" + ",".join(f"{L(f'Sem {k}')}{lin}" for k in range(1, N_SEM + 1)) + ")",
        }
        # alvo semanal da pessoa (CADASTRO, coluna Alvo/sem — pelo nome) menos o
        # desconto por dia de ausência na semana (CONFIG). BH = horas − alvo.
        alvo = (f'IFERROR(INDEX(CADASTRO!$I$4:$I$80,MATCH($A{lin},CADASTRO!$A$4:$A$80,0)),'
                f'${LC}{lin})')
        for k in range(1, N_SEM + 1):
            if k <= len(jans):
                spans = jans[k - 1]
                formulas[f"Sem {k}"] = "=" + "+".join(
                    f"SUMPRODUCT({expr_horas(lin, a2, b2)})" for a2, b2 in spans)
                formulas[f"BH {k}"] = (f"={L(f'Sem {k}')}{lin}-MAX(0,{alvo}"
                                       f"-CONFIG!$B${r_desc}*({expr_ausencias(lin, spans)}))")
            else:
                formulas[f"Sem {k}"] = formulas[f"BH {k}"] = ""
        for nome in COLS_TOT:
            if nome in ("Grupo", "Nº"):
                continue
            formula = formulas[nome]
            c = ws.cell(row=lin, column=col_tot(nome), value=formula or None)
            c.number_format = "+0;-0;0" if nome.startswith("BH ") else "0"
            c.font = Font(name=F, size=8, bold=(nome in ("Saldo",) or nome.startswith("BH ")),
                          color=INK if (nome in COLS_ALERTA or nome == "Saldo"
                                        or nome.startswith("BH ")) else INK2)
            c.alignment = Alignment(horizontal="center")
            c.border = BOX
        cg = ws.cell(row=lin, column=C_TOT + IDX_GRUPO, value=grupo_filtro(apelido))
        cg.font = Font(name=F, size=8, color=INK3)
        cg.border = BOX
        cn = ws.cell(row=lin, column=C_TOT + IDX_NUM, value=k + 1)
        cn.font = Font(name=F, size=8, color=INK3)
        cn.alignment = Alignment(horizontal="center")
        cn.border = BOX

    ultima = R_P0 + len(pessoas) - 1
    # ------- rodapé: lotação e falta
    r = ultima + 2
    # os grupos derivam dos flags da tabela de turnos (CONFIG colunas "conta ...").
    # Não repetir a lista à mão: foi assim que P/R entraram na cobertura sem querer.
    grupos = [(rot, [k for k, v in D.TURNOS.items() if v[4 + i]], i)
              for i, rot in enumerate(("Manhã", "Tarde", "Noite"))]
    linhas_falta = []
    for gi, (rot, letras, idx) in enumerate(grupos):
        rl = r + gi
        cl = ws.cell(row=rl, column=C_MED, value=rot)
        cl.font = Font(name=F, size=9, bold=True, color=LAVI)
        _tip(cl, "cob-turno")
        rf = r + 3 + gi
        cf0 = ws.cell(row=rf, column=C_MED, value=f"Falta {rot.lower()}")
        cf0.font = Font(name=F, size=9, bold=True, color=CORALI)
        _tip(cf0, "falta")
        linhas_falta.append(rf)
        # vésperas entram na contagem visual (a semana inteira à vista);
        # o dashboard só varre C_D1..C_DN, então elas não poluem a média do mês
        vesperas = ([(-k, dt.date(2026, mes, 1) - dt.timedelta(days=k))
                     for k in range(offset, 0, -1)] if aba_prev else [])
        virada = ([(31 + j, dias_virada[j]) for j in range(n_virada)]
                  if aba_prox else [])
        for i, data in vesperas + \
                [(i2, dt.date(2026, mes, i2 + 1)) for i2 in range(ndias)] + virada:
            col = get_column_letter(C_D1 + i)
            faixa = f"{col}{R_P0}:{col}{ultima}"
            conta = "+".join(f'COUNTIF({faixa},"{x}")' for x in letras)
            c = ws.cell(row=rl, column=C_D1 + i, value=f"={conta}")
            c.font = Font(name=F, size=8, color=INK2)
            c.alignment = Alignment(horizontal="center")
            # feriado escala como o dia da semana em que cai; véspera e dia 32
            # usam a vigência do mês a que pertencem (SET→01/10 já é regra nova)
            mes_do_dia = (mes - 1) if i < 0 else (mes if i < 31 else mes % 12 + 1)
            base = 4 if mes_do_dia >= D.VIGENCIA_NOVA else 9
            tipo = base + (0 if data.weekday() < 5 else (1 if data.weekday() == 5 else 2))
            ref = f"CONFIG!${get_column_letter(2+idx)}${tipo}"
            cf = ws.cell(row=rf, column=C_D1 + i,
                         value=f"=MAX(0,{ref}-{col}{rl})")
            cf.font = Font(name=F, size=8, bold=True, color=CORALI)
            cf.alignment = Alignment(horizontal="center")
    # Sem⚠ e as semanas civis: vermelho acima do teto constitucional de 44h
    for nome in [f"Sem {k}" for k in range(1, N_SEM + 1)] + ["Sem⚠"]:
        lc = get_column_letter(col_tot(nome))
        ws.conditional_formatting.add(
            f"{lc}{R_P0}:{lc}{ultima}",
            CellIsRule(operator="greaterThan", formula=["44"],
                       fill=PatternFill("solid", bgColor=CORAL),
                       font=Font(name=F, size=8, bold=True, color="FFFFFF")))
    # BH da semana: positivo = BHP (areia) · negativo = BHN / faltou (coral claro)
    for k in range(1, N_SEM + 1):
        lc = get_column_letter(col_tot(f"BH {k}"))
        faixa_bh = f"{lc}{R_P0}:{lc}{ultima}"
        ws.conditional_formatting.add(faixa_bh, CellIsRule(
            operator="greaterThan", formula=["0"],
            fill=PatternFill("solid", bgColor=SAND),
            font=Font(name=F, size=8, bold=True, color=INK)))
        ws.conditional_formatting.add(faixa_bh, CellIsRule(
            operator="lessThan", formula=["0"],
            fill=PatternFill("solid", bgColor=CORALS),
            font=Font(name=F, size=8, bold=True, color=CORALI)))
    col_exc = get_column_letter(col_tot("FDS⚠"))
    ws.conditional_formatting.add(
        f"{col_exc}{R_P0}:{col_exc}{ultima}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", bgColor=CORAL),
                   font=Font(name=F, size=8, bold=True, color="FFFFFF")))
    for rf in linhas_falta:
        faixa = f"{L_PRE1}{rf}:{get_column_letter(C_D32 + N_POS - 1)}{rf}"
        ws.conditional_formatting.add(faixa, CellIsRule(
            operator="greaterThan", formula=["0"],
            fill=PatternFill("solid", bgColor=CORAL),
            font=Font(name=F, size=8, bold=True, color="FFFFFF")))
    creme(ws, linhas_falta[-1] + 2, C_TOT + len(COLS_TOT))
    ws.freeze_panes = f"{L_PRE1}{R_P0}"
    return ws


def aba_painel(wb, pessoas, oficial):
    ws = wb.create_sheet("PAINEL ANO")
    ws.sheet_properties.tabColor = LAVI
    estilo_titulo(ws, "Painel do ano · O comparativo",
                  "Cada número vem da aba do mês — não se digita nada aqui")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 5
    ws.cell(row=R_HDR, column=C_MED, value="Médico").font = Font(
        name=F, bold=True, size=9, color="FFFFFF")
    ws.cell(row=R_HDR, column=C_MED).fill = PatternFill("solid", fgColor=LAVI)
    ws.cell(row=R_HDR, column=C_CH, value="CH").font = Font(
        name=F, bold=True, size=9, color="FFFFFF")
    ws.cell(row=R_HDR, column=C_CH).fill = PatternFill("solid", fgColor=LAVI)

    # 4 blocos de 13 colunas (12 meses + total)
    blocos = [("Saldo do mês (h)", col_tot("Saldo"), LAVS),
              ("FDS (h)", col_tot("FDS"), AQUAS),
              ("Sexta-noite", col_tot("SxN"), SANDS),
              ("Feriado (h)", col_tot("Feriado"), CORALS)]
    chave_bloco = {"Saldo do mês (h)": "saldo-bloco", "FDS (h)": "fds-bloco",
                   "Sexta-noite": "sxn-bloco", "Feriado (h)": "feriado-bloco"}
    col = 3
    inicio_bloco = {}
    for rot, col_origem, cor in blocos:
        cb = ws.cell(row=R_DOW, column=col, value=rot)
        cb.font = Font(name=F, bold=True, size=10, color=LAVI)
        _tip(cb, chave_bloco.get(rot, ""))
        ws.merge_cells(start_row=R_DOW, start_column=col, end_row=R_DOW, end_column=col + 12)
        inicio_bloco[rot] = col
        for i, m in enumerate(MESES_PT):
            c = ws.cell(row=R_HDR, column=col + i, value=m)
            c.font = Font(name=F, bold=True, size=8, color=INK)
            c.fill = PatternFill("solid", fgColor=cor)
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col + i)].width = 5.4
        c = ws.cell(row=R_HDR, column=col + 12, value="Ano")
        c.font = Font(name=F, bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=LAVI)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col + 12)].width = 6.5
        for k in range(len(pessoas)):
            lin = R_P0 + k
            for i, m in enumerate(MESES_PT):
                # pelo NOME, não pela posição: a aba do mês pode estar ordenada
                # pelo filtro e o painel continua achando cada pessoa
                lc = get_column_letter(col_origem)
                cel = ws.cell(row=lin, column=col + i, value=(
                    f"=IFERROR(INDEX({m}!${lc}${R_P0}:${lc}${R_P0+65},"
                    f"MATCH($A{lin},{m}!$A${R_P0}:$A${R_P0+65},0)),\"\")"))
                cel.number_format = "0"
                cel.font = Font(name=F, size=8, color=INK2)
                cel.alignment = Alignment(horizontal="center")
            tot = ws.cell(row=lin, column=col + 12,
                          value=f"=SUM({get_column_letter(col)}{lin}:"
                                f"{get_column_letter(col+11)}{lin})")
            tot.number_format = "0"
            tot.font = Font(name=F, size=8, bold=True, color=INK)
            tot.alignment = Alignment(horizontal="center")
            tot.fill = PatternFill("solid", fgColor=CREME)
        col += 14

    # conferência com a contagem manual antiga
    ws.cell(row=R_DOW, column=col, value="Conferência · contagem manual antiga").font = Font(
        name=F, bold=True, size=10, color=CORALI)
    ws.merge_cells(start_row=R_DOW, start_column=col, end_row=R_DOW, end_column=col + 3)
    chave_conf = {"SxN oficial": "sxn-oficial", "SxN calc": "sxn-calc", "difere?": "difere"}
    for i, h in enumerate(("SxN oficial", "SxN calc", "difere?", "")):
        if not h:
            continue
        c = _tip(ws.cell(row=R_HDR, column=col + i, value=h), chave_conf.get(h, ""))
        c.font = Font(name=F, bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=CORALI)
        c.alignment = VERTICAL
        ws.column_dimensions[get_column_letter(col + i)].width = 5.4
    ws.row_dimensions[R_HDR].height = 58

    for k, (apelido, _ch) in enumerate(pessoas):
        lin = R_P0 + k
        c = ws.cell(row=lin, column=C_MED, value=apelido)
        c.font = Font(name=F, size=9, bold=True, color=INK)
        ws.cell(row=lin, column=C_CH,
                value=f"=CADASTRO!$C{4+k}").font = Font(name=F, size=8, color=INK3)
        tem_oficial = any(n == apelido for n, _m in oficial)
        sxn_of = sum(v for (n, _m), v in oficial.items() if n == apelido)
        cel = ws.cell(row=lin, column=col, value=sxn_of if tem_oficial else None)
        cel.font = Font(name=F, size=8, color=INK3)
        cel.alignment = Alignment(horizontal="center")
        base = get_column_letter(inicio_bloco["Sexta-noite"])
        fim = get_column_letter(inicio_bloco["Sexta-noite"] + 11)
        cc = ws.cell(row=lin, column=col + 1, value=f"=SUM({base}{lin}:{fim}{lin})")
        cc.font = Font(name=F, size=8, color=INK2)
        cc.alignment = Alignment(horizontal="center")
        L = get_column_letter(col)
        cd = ws.cell(row=lin, column=col + 2,
                     value=f'=IF({L}{lin}="","",IF({L}{lin}={get_column_letter(col+1)}{lin},"","⚠"))')
        cd.font = Font(name=F, size=9, bold=True, color=CORALI)
        cd.alignment = Alignment(horizontal="center")
    creme(ws, R_P0 + len(pessoas) + 1, col + 4)
    ws.freeze_panes = "C8"
    return ws


def aba_senior(wb, pessoas, fim_cod):
    ws = wb.create_sheet("SENIOR")
    ws.sheet_properties.tabColor = SAND
    ws.sheet_view.showGridLines = False
    t = ws.cell(row=R_TIT, column=1, value="Códigos Senior")
    t.font = Font(name=DISPLAY, bold=True, size=14, color=INK)
    sel = ws.cell(row=R_TIT, column=C_CH, value="OUT")
    sel.font = Font(name=F, bold=True, size=13, color=CORALI)
    sel.fill = PatternFill("solid", fgColor=SANDS)
    sel.alignment = Alignment(horizontal="center")
    dv = DataValidation(type="list", formula1='"' + ",".join(MESES_PT) + '"',
                        allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(sel)
    inst = ws.cell(row=R_TIT, column=3,
                   value="Escolha o mês na célula ao lado — a matriz inteira se traduz "
                         "nos códigos do RH, pronta para lançar no Senior. BHP (+) não "
                         "gera código; BHN (−) sai como o plantão normal")
    inst.font = Font(name=F, size=9, italic=True, color=INK3)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 6
    for col, txt in ((C_MED, "Médico"), (C_CH, "Apelido")):
        c = ws.cell(row=R_HDR, column=col, value=txt)
        c.font = Font(name=F, bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=LAVI)
    # a matriz aqui é compacta (dia 1º na coluna C): as vésperas do mês anterior
    # NÃO entram — código do RH se lança uma vez, no mês dono do dia. Por isso o
    # salto de N_PRE colunas ao mirar a aba mensal.
    S_D1 = 3
    for i in range(32):
        col = S_D1 + i
        alvo = get_column_letter(col + N_PRE)
        ws.column_dimensions[get_column_letter(col)].width = 6.5
        for r, formula in ((R_DIA, f'=IFERROR(INDIRECT($B$1&"!{alvo}{R_DIA}"),"")'),
                           (R_DOW, f'=IFERROR(INDIRECT($B$1&"!{alvo}{R_DOW}"),"")')):
            c = ws.cell(row=r, column=col, value=formula)
            c.font = Font(name=F, bold=(r == R_DIA), size=8, color=INK if r == R_DIA else INK3)
            c.alignment = Alignment(horizontal="center")

    nomes = {a: n for a, n, *_ in D.ROSTER}
    faixa_cod = f"CONFIG!$A$15:$B${fim_cod}"
    for k, (apelido, _ch) in enumerate(pessoas):
        lin = R_P0 + k
        c = ws.cell(row=lin, column=C_MED, value=nomes.get(apelido, apelido))
        c.font = Font(name=F, size=9, color=INK)
        # apelido na coluna B: é por ele que cada célula ACHA a pessoa na aba
        # do mês (MATCH pelo nome — a aba pode estar ordenada pelo filtro)
        cb = ws.cell(row=lin, column=2, value=apelido)
        cb.font = Font(name=F, size=7, color=INK3)
        for i in range(32):
            col = S_D1 + i
            cel = ws.cell(row=lin, column=col, value=(
                f'=IFERROR(VLOOKUP(INDEX('
                f'INDIRECT($B$1&"!"&ADDRESS({R_P0},COLUMN()+{N_PRE},4)&":"'
                f'&ADDRESS({R_P0+65},COLUMN()+{N_PRE},4)),'
                f'MATCH($B{lin},INDIRECT($B$1&"!$A${R_P0}:$A${R_P0+65}"),0)),'
                f'{faixa_cod},2,FALSE)&"","")'))       # &"": código vazio (BHP) fica vazio, não 0
            cel.font = Font(name=F, size=8, color=INK)
            cel.alignment = Alignment(horizontal="center")
            cel.border = BOX
    creme(ws, R_P0 + len(pessoas) + 1, S_D1 + 32)
    ws.freeze_panes = f"{get_column_letter(S_D1)}{R_P0}"
    return ws


def aba_dia_a_dia(wb, DIAS, mes=10):
    """o mês em formato calendário: por dia, por turno, quem está onde.

    Tudo FÓRMULA sobre a aba do mês (FILTER + TEXTJOIN, que existem no Google
    Sheets e no Excel 365): trocou um plantão no dropdown da matriz, este
    calendário se refaz. Serve pra "perceber o que tá esquisito" — pedido do
    Marcos em 18/08/26.
    """
    nome_mes = MESES_PT[mes - 1]
    ws = wb.create_sheet(f"{nome_mes} · DIA A DIA",
                         wb.sheetnames.index(nome_mes) + 1)
    ws.sheet_properties.tabColor = CORALI
    ws.sheet_view.showGridLines = False
    t = ws.cell(row=1, column=1, value=f"{MESES_LONGO[mes-1].capitalize()} · Dia a dia")
    t.font = Font(name=DISPLAY, bold=True, size=14, color=INK)
    sub = ws.cell(row=1, column=4, value=(
        "Cada célula lista quem está no turno — é fórmula da aba do mês: "
        "mudou lá, muda aqui na hora"))
    sub.font = Font(name=F, size=9, italic=True, color=INK3)
    ndias = (dt.date(2026, mes % 12 + 1, 1) - dt.timedelta(days=1)).day

    # sem coluna de ausências (pedido do Marcos, 28/08/26): quem está fora
    # já se percebe pela própria matriz do mês — aqui é só quem trabalha
    colunas = [("Dia", 5), ("", 5), ("Manhã", 44), ("Tarde", 36), ("Noite", 30),
               ("BHN (dispensa)", 22), ("Cobertura", 13)]
    chave_cal = {"Cobertura": "cobertura-cal", "BHN (dispensa)": "bhn-cal"}
    for i, (h, w) in enumerate(colunas, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name=F, bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=LAVI)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
        if h in chave_cal:
            _tip(c, chave_cal[h])

    nomes_rng = f"{nome_mes}!$A${R_P0}:$A${R_P0+65}"

    def formula_turno(col, letras, bhp=()):
        """nomes do turno; quem está de BHP naquele período ganha ' BHP' colado
        (o formato da grade oficial — é assim que a Mari confere)."""
        rng = f"{nome_mes}!{col}${R_P0}:{col}${R_P0+65}"
        cond = "+".join(f'({rng}="{l}")' for l in letras)
        rotulo = nomes_rng
        if bhp:
            cond_bhp = "+".join(f'({rng}="{l}")' for l in bhp)
            rotulo = f'{nomes_rng}&IF(({cond_bhp})*1>0," BHP","")'
        return (f'=IFERROR(TEXTJOIN(", ",TRUE,'
                f'FILTER({rotulo},({cond})*1>0)),"—")')

    def formula_bhn(col):
        """a linha 'lá embaixo' da grade oficial: quem foi dispensado (BHN) no dia."""
        rng = f"{nome_mes}!{col}${R_P0}:{col}${R_P0+65}"
        cond = "+".join(f'({rng}="{l}")' for l in sorted(D.BHN))
        periodo = (f'IF({rng}="M-"," manhã",IF({rng}="T-"," tarde",IF({rng}="D-"," dia",'
                   f'IF({rng}="N-"," noite",IF({rng}="Tm-"," manhã"," tarde")))))')
        return (f'=IFERROR(TEXTJOIN(", ",TRUE,'
                f'FILTER({nomes_rng}&" BHN"&{periodo},({cond})*1>0)),"")')

    # a semana inteira, de segunda a domingo, nas duas pontas (Marcos, 28/08):
    # o calendário abre na segunda da semana do dia 1º (vésperas) e fecha no
    # domingo da última semana (virada) — mesmas colunas vivas da aba do mês
    offset = dt.date(2026, mes, 1).weekday()
    n_virada = (7 - ((offset + ndias) % 7)) % 7
    dias = []
    if mes > 1:
        for j in range(offset):
            data = dt.date(2026, mes, 1) - dt.timedelta(days=offset - j)
            dias.append((data, C_D1 - offset + j, True))
    dias += [(dt.date(2026, mes, d + 1), C_D1 + d, False) for d in range(ndias)]
    if mes < 12:
        prox1 = dt.date(2026, mes % 12 + 1, 1)
        for j in range(n_virada):
            dias.append((prox1 + dt.timedelta(days=j), C_D32 + j, True))

    r = 3
    for data, col_n, vizinho in dias:
        r += 1
        fds = data.weekday() >= 5
        info_fer = D.FERIADOS_2026.get((data.month, data.day))
        cd = ws.cell(row=r, column=1,
                     value=data.strftime("%d/%m") if vizinho else data.day)
        cd.font = Font(name=F, bold=not vizinho, size=8 if vizinho else 11,
                       color="FFFFFF" if info_fer else (INK3 if vizinho else INK))
        cd.alignment = Alignment(horizontal="center", vertical="top")
        rotulo = info_fer[1] if info_fer else WD_PT[data.weekday()]
        cw = ws.cell(row=r, column=2, value=rotulo)
        cw.font = Font(name=F, size=9, bold=bool(info_fer),
                       color="FFFFFF" if info_fer else (CORALI if fds else INK3))
        cw.alignment = Alignment(horizontal="center", vertical="top")
        if info_fer:
            cd.fill = PatternFill("solid", fgColor=CORAL)
            cw.fill = PatternFill("solid", fgColor=CORALI)
        elif fds:
            for cc in (cd, cw):
                cc.fill = PatternFill("solid", fgColor=LAVS)
        col = get_column_letter(col_n)
        # os códigos de cada turno vêm dos flags da tabela (conta manhã/tarde/noite)
        grupos_cal = ((3, [k for k, v in D.TURNOS.items() if v[4]], ("M+", "D+", "Dm+")),
                      (4, [k for k, v in D.TURNOS.items() if v[5]], ("T+", "D+", "Dt+")),
                      (5, [k for k, v in D.TURNOS.items() if v[6]], ("N+",)))
        for j, letras, bhp in grupos_cal:
            cel = ws.cell(row=r, column=j, value=formula_turno(col, letras, bhp))
            cel.font = Font(name=F, size=8, color=INK3 if vizinho else INK,
                            italic=vizinho)
            cel.alignment = Alignment(wrap_text=True, vertical="top")
            cel.border = BOX
            if vizinho or fds or info_fer:
                cel.fill = PatternFill("solid", fgColor=CREME2)
        cb = ws.cell(row=r, column=6, value=formula_bhn(col))
        cb.font = Font(name=F, size=8, italic=True, color=INK3)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        cb.border = BOX
        cb.fill = PatternFill("solid", fgColor=FILL_BHN)
        # "M18 T12 N9 ✓" — sem denominador enganoso: quando sobra gente, o
        # x/(x+falta) da versão anterior mostrava "18/18" como se o alvo fosse 18
        f_falta = (f'{nome_mes}!{col}${R_P0+70}+{nome_mes}!{col}${R_P0+71}'
                   f'+{nome_mes}!{col}${R_P0+72}')
        cov = ws.cell(row=r, column=7, value=(
            f'="M"&{nome_mes}!{col}${R_P0+67}&" T"&{nome_mes}!{col}${R_P0+68}'
            f'&" N"&{nome_mes}!{col}${R_P0+69}'
            f'&IF({f_falta}=0," ✓"," ⚠ falta "&({f_falta}))'))
        cov.font = Font(name=F, size=8, color=INK3 if vizinho else INK2,
                        italic=vizinho)
        cov.alignment = Alignment(wrap_text=True, vertical="top")
        cov.border = BOX
        # altura pelo que vai aparecer: quantos nomes cada turno tem naquele dia
        import math
        pessoas_dia = DIAS.get(data, {})
        def chars(flag):
            return sum(len(p) + 2 + (4 if cel[0] in D.BHP else 0)
                       for p, cel in pessoas_dia.items()
                       if cel[0] in D.TURNOS and D.TURNOS[cel[0]][flag])
        bhn = sum(len(p) + 12 for p, cel in pessoas_dia.items() if cel[0] in D.BHN)
        linhas = max(math.ceil(chars(4) / (44 * 1.35)), math.ceil(chars(5) / (36 * 1.35)),
                     math.ceil(chars(6) / (30 * 1.35)), math.ceil(bhn / (22 * 1.35)), 1)
        ws.row_dimensions[r].height = max(26, linhas * 11.6 + 8)
    creme(ws, r + 2, 8)
    ws.freeze_panes = "C4"
    return ws


def contagem_oficial_sxn():
    """lê a aba SEXTA NOITE da contagem anual da Mari: {(apelido, mês): nº}."""
    import unicodedata
    caminho = os.path.join(senior_import.DIR_FONTES,
                           "CONTAGEM FDS - SEXTA NOITE - FERIADO 2026"
                           "(Recuperado Automaticamente).xlsx")
    if not os.path.exists(caminho):
        return {}
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["SEXTA NOITE"]
    alias = {"Mpinheiro": "MPinheiro", "Letícia": "Leticia", "PJamile": "Pjamile",
             "Raquel Assis": "Raquel", "Marilia": "Marilia"}
    colunas = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str):
            up = unicodedata.normalize("NFC", v.strip()).upper()
            if up in senior_import.MESES:
                colunas[senior_import.MESES[up]] = c
    saida = {}
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(row=r, column=2).value
        if not nome or not str(nome).strip():
            continue
        nome = unicodedata.normalize("NFC", str(nome).strip())
        nome = alias.get(nome, nome)
        for m, c in colunas.items():
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                saida[(nome, m)] = saida.get((nome, m), 0) + int(v)
    return saida


def aba_validador(wb, DIAS, pessoas):
    ws = wb.create_sheet("VALIDADOR")
    ws.sheet_properties.tabColor = CORALI
    estilo_titulo(ws, "Validador · O que fere regra dura",
                  "Estrutural = a forma do contrato · Pontual = erro daquele mês")
    for col, w in zip("ABCDEF", (15, 13, 13, 62, 17, 30)):
        ws.column_dimensions[col].width = w

    grupo_de = {a: g for a, _n, _c, g, *_ in D.ROSTER}
    achados = auditar(DIAS, [p for p, _ in pessoas])
    from collections import Counter
    dsr_por_pessoa = Counter(a["pessoa"] for a in achados if a["tipo"] == "DSR")

    estrutural, alerta, registro = [], [], []
    for a in achados:
        if a["tipo"] == "DSR":
            if grupo_de.get(a["pessoa"]) in ("chefia", "rotina", "administrativo") \
                    or dsr_por_pessoa[a["pessoa"]] >= 3:
                estrutural.append(a)
            else:
                alerta.append(a)
        elif a["horas"] <= 2:
            alerta.append(a)
        else:
            registro.append(a)

    r = 3

    def secao(titulo, cor, explicacao):
        nonlocal r
        c = ws.cell(row=r, column=1, value=titulo)
        c.font = Font(name=DISPLAY, bold=True, size=12, color=cor)
        r += 1
        e = ws.cell(row=r, column=1, value=explicacao)
        e.font = Font(name=F, size=9, italic=True, color=INK2)
        e.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.row_dimensions[r].height = altura(explicacao, 150)
        r += 2

    def cabecalho():
        nonlocal r
        for i, h in enumerate(("Médico", "Grupo", "Quando", "O que acontece",
                               "Base", "Tratamento"), start=1):
            c = ws.cell(row=r, column=i, value=h)
            c.font = Font(name=F, bold=True, size=8, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=LAVI)
        r += 1

    def despejar(itens, trat, cor):
        nonlocal r
        for a in itens:
            ws.cell(row=r, column=1, value=a["pessoa"]).font = Font(
                name=F, size=9, bold=True, color=INK)
            ws.cell(row=r, column=2, value=grupo_de.get(a["pessoa"], "—")).font = Font(
                name=F, size=8, color=INK3)
            ws.cell(row=r, column=3, value=f'{a["data"]:%d/%m}').font = Font(
                name=F, size=8, color=INK2)
            cd = ws.cell(row=r, column=4, value=a["detalhe"])
            cd.font = Font(name=F, size=9, color=INK)
            cd.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=5, value=a["regra"]).font = Font(
                name=F, size=8, color=INK3)
            ct = ws.cell(row=r, column=6, value=trat)
            ct.font = Font(name=F, size=8, bold=True, color=cor)
            for i in range(1, 7):
                ws.cell(row=r, column=i).border = BOX
                if i != 4:
                    ws.cell(row=r, column=i).alignment = Alignment(vertical="top")
            ws.row_dimensions[r].height = altura(a["detalhe"], 62, minimo=15)
            r += 1

    secao("ESTRUTURAL — decisão de política, não erro de montagem", LAVI,
          "Quem faz 6h de manhã todos os dias tem folga de 18h entre turnos (13h → 07h) e "
          "nunca alcança as 24h consecutivas do art. 67. Não é descuido da escala: é a forma "
          "do contrato de diarista. Aparece aqui uma vez por semana afetada, pra ficar "
          "visível sem virar alarme. Resolver isso é decisão da chefia, não da escalista.")
    cabecalho()
    por_pessoa = {}
    for a in estrutural:
        por_pessoa.setdefault(a["pessoa"], []).append(a)
    for pessoa, itens in sorted(por_pessoa.items(), key=lambda x: -len(x[1])):
        ws.cell(row=r, column=1, value=pessoa).font = Font(name=F, size=9, bold=True, color=INK)
        ws.cell(row=r, column=2, value=grupo_de.get(pessoa, "—")).font = Font(
            name=F, size=8, color=INK3)
        ws.cell(row=r, column=3, value=f"{len(itens)} semanas").font = Font(
            name=F, size=8, color=INK2)
        cd = ws.cell(row=r, column=4, value="sem folga de 24h consecutivas nessas semanas: "
                     + ", ".join(f'{a["data"]:%d/%m}' for a in itens))
        cd.font = Font(name=F, size=9, color=INK)
        cd.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=5, value="art. 67 CLT").font = Font(name=F, size=8, color=INK3)
        ws.cell(row=r, column=6, value="ESTRUTURAL").font = Font(
            name=F, size=8, bold=True, color=LAVI)
        for i in range(1, 7):
            ws.cell(row=r, column=i).border = BOX
        ws.row_dimensions[r].height = altura(cd.value, 62, minimo=18)
        r += 1
    r += 2

    secao(f"ALERTA — pontual, o alvo é zero ({len(alerta)})", CORALI,
          "Descanso de 2h ou menos entre jornadas: na prática é jornada de 18h emendada. "
          "Era ~15 casos por mês no 1º trimestre de 2026 e caiu pra ~2 depois da mudança de "
          "abril — a regra apertou e funcionou. O papel desta aba é segurar a linha no zero. "
          "Nas abas mensais isto é a coluna 18h⚠, que recalcula sozinha ao digitar.")
    cabecalho()
    despejar(alerta, "ALERTA", CORALI)
    r += 2

    secao(f"REGISTRO — prática estabelecida do serviço ({len(registro)})", INK2,
          "Noite 19–07h seguida da tarde 13–19h: 18h de trabalho com 6h de pausa. São 90 "
          "casos em 6 meses, 26 pessoas, sem queda no ano — é como o serviço funciona, não "
          "acidente. Decisão de 17/08/26: continua e fica registrado, sem alarme. Fica aqui "
          "porque está abaixo das 11h do art. 66 e, se algum dia for questionado, o registro "
          "existe. Observação honesta: o arquivo de códigos não carrega BHP/BHN nem troca, "
          "então parte destes casos pode ser troca documentada.")
    cabecalho()
    despejar(registro, "REGISTRO", INK2)
    r += 2

    # adicional noturno
    c = ws.cell(row=r, column=1, value="ADICIONAL NOTURNO — horas em 22h–05h, por pessoa e mês")
    c.font = Font(name=DISPLAY, bold=True, size=12, color=LAVI)
    r += 1
    e = ws.cell(row=r, column=1, value="Insumo da folha: a hora noturna vale 52min30s e o "
                "adicional é de no mínimo 20% (art. 73 CLT). A noite de 19–07h cai inteira "
                "dentro da janela nas 7 horas entre 22h e 05h.")
    e.font = Font(name=F, size=9, italic=True, color=INK2)
    e.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = altura(e.value, 150)
    r += 2
    noturnas = noturnas_por_mes(DIAS, [p for p, _ in pessoas])
    ws.cell(row=r, column=1, value="Médico").font = Font(name=F, bold=True, size=8, color="FFFFFF")
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LAVI)
    for i, m in enumerate(MESES_PT):
        cc = ws.cell(row=r, column=2 + i, value=m)
        cc.font = Font(name=F, bold=True, size=8, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor=LAVI)
        cc.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=14, value="Ano").font = Font(name=F, bold=True, size=8, color="FFFFFF")
    ws.cell(row=r, column=14).fill = PatternFill("solid", fgColor=CORALI)
    r += 1
    for pessoa, _ch in pessoas:
        total = 0
        tem = False
        ws.cell(row=r, column=1, value=pessoa).font = Font(name=F, size=9, bold=True, color=INK)
        for i, m in enumerate(range(1, 13)):
            v = noturnas.get((pessoa, m))
            if v:
                tem = True
                total += v
                cc = ws.cell(row=r, column=2 + i, value=round(v, 1))
                cc.font = Font(name=F, size=8, color=INK2)
                cc.alignment = Alignment(horizontal="center")
        ct = ws.cell(row=r, column=14, value=round(total, 1) if total else None)
        ct.font = Font(name=F, size=8, bold=True, color=INK)
        ct.alignment = Alignment(horizontal="center")
        if tem:
            r += 1
        else:
            for i in range(1, 15):
                ws.cell(row=r, column=i).value = None
    creme(ws, r + 2, 14)
    ws.freeze_panes = "A4"
    return ws


def aba_checagem(wb):
    """a checagem de outubro (e-mail Marcos/Mari): regras reveladas, o que foi
    mudado célula a célula e o que foi verificado sem mudar — pra Mari conferir."""
    import checagem_out_v3 as CK
    ws = wb.create_sheet("CHECAGEM OUT", wb.sheetnames.index("OUT · DIA A DIA") + 1)
    ws.sheet_properties.tabColor = CORALI
    ws.sheet_view.showGridLines = False
    estilo_titulo(ws, "Checagem de outubro · O que mudou e por quê",
                  "Correções do e-mail de 28/08 aplicadas sobre a versão viva da Mari (01/09)")
    for col, w in zip("ABCDEF", (6, 13, 7, 8, 8, 110)):
        ws.column_dimensions[col].width = w
    r = 4

    def titulo(txt):
        nonlocal r
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(name=DISPLAY, bold=True, size=12, color=LAVI)
        r += 1

    def cab(textos):
        nonlocal r
        for i, t in enumerate(textos, start=1):
            c = ws.cell(row=r, column=i, value=t)
            c.font = Font(name=F, bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=LAVI)
        r += 1

    titulo("As regras que a checagem revelou (e a planilha passou a medir)")
    cab(["", "Regra", "", "", "", "O que é"])
    for nome, txt in CK.REGRAS:
        cn = ws.cell(row=r, column=2, value=nome)
        cn.font = Font(name=F, size=9, bold=True, color=INK)
        cn.alignment = Alignment(vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        c = ws.cell(row=r, column=6, value=txt)
        c.font = Font(name=F, size=9, color=INK2)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        for j in range(2, 7):
            ws.cell(row=r, column=j).border = BOX
        ws.row_dimensions[r].height = altura(txt, 110)
        r += 1
    r += 1
    titulo("Célula a célula — o que mudou na aba OUT")
    cab(["Item", "Médico", "Dia", "Estava", "Ficou", "Por quê"])
    for item, ap, dia, de, para, motivo in CK.EDITS:
        vals = [item, ap, "01/11" if dia == 32 else f"{dia:02d}/10", de or "—", para or "—", motivo]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = Font(name=F, size=9, color=INK, bold=(j in (2, 5)))
            c.alignment = Alignment(horizontal="left" if j == 6 else "center",
                                    wrap_text=(j == 6), vertical="top")
            c.border = BOX
        if para in D.BHP:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=FILL_BHP)
        elif para in D.BHN:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=FILL_BHN)
        ws.row_dimensions[r].height = altura(motivo, 110, minimo=18)
        r += 1
    r += 1
    titulo("Verificado sem mudar célula")
    cab(["Item", "Quem", "", "", "", "O que foi visto"])
    for item, ap, txt in CK.SEM_ALTERACAO:
        ws.cell(row=r, column=1, value=item).font = Font(name=F, size=9, color=INK)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ca = ws.cell(row=r, column=2, value=ap)
        ca.font = Font(name=F, size=9, bold=True, color=INK)
        ca.alignment = Alignment(vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        c = ws.cell(row=r, column=6, value=txt)
        c.font = Font(name=F, size=9, color=INK2)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        for j in range(1, 7):
            ws.cell(row=r, column=j).border = BOX
        ws.row_dimensions[r].height = altura(txt, 110)
        r += 1
    creme(ws, r + 2, 6)
    ws.freeze_panes = "A4"
    return ws


NOME_ARQUIVO = "Escala UTI HCB 2026 - V3.xlsx"


def main():
    DIAS, rel_grade = carregar_dados()
    oficial = contagem_oficial_sxn()
    pessoas = [(a, ch) for a, _n, ch, *_ in D.ROSTER]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    aba_leiame(wb, rel_grade)
    aba_cadastro(wb)
    _cfg, fim_cod, r_cota, r_desc = aba_config(wb)
    for mes in range(1, 13):
        aba_mes(wb, mes, DIAS, fim_cod, pessoas, r_cota, r_desc)
    aba_dia_a_dia(wb, DIAS, mes=10)
    if MODO == "v3":
        aba_checagem(wb)
    aba_painel(wb, pessoas, oficial)
    aba_senior(wb, pessoas, fim_cod)
    aba_validador(wb, DIAS, pessoas)
    # painel de leitura: DADOS DASH alimenta os gráficos, DASHBOARD é a capa
    import v4_dashboard as VD
    VD.aba_dados(wb)
    VD.aba_dashboard(wb, mes_vivo="OUT", nome_mes="outubro", nota=(
        None if MODO == "v3" else
        "Outubro é o que está lançado nesta planilha (lido ao vivo em 02/09). A estrutura é a da "
        "V3 — semanas e BH na frente, códigos de banco, cabeçalho vertical — sem nenhuma "
        "correção da checagem: essas estão só na V3."))

    destino = os.path.join(AQUI, NOME_ARQUIVO if MODO == "v3"
                           else "Escala UTI HCB 2026 - unificada v4.xlsx")
    wb.save(destino)
    tam = os.path.getsize(destino) / 1024
    print(f"\n=== planilha V3 salva ===\n{destino}\n{tam:.0f} kb · {len(wb.sheetnames)} abas")
    print("abas:", " · ".join(wb.sheetnames))
    return destino


if __name__ == "__main__":
    main()
