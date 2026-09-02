#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Importa jan/mai/jun/set de 2026 a partir das GRADES do grupo.

A grade é um bloco por dia com os nomes empilhados em colunas Manhã | Tarde | Noite | Noitinha.
Reconstrução: manhã+tarde no mesmo dia = 12h dia (D); só manhã = M; só tarde = T;
noite = N; noitinha = NT.

FIDELIDADE MENOR que os meses vindos dos códigos Senior — registrar isso na planilha:
- 47 (10h chefia) aparece só como nome na coluna Manhã → entra como M (6h), perde 4h;
- 6 (4h CEP) e 78 (5h Janaina) idem, viram M/T;
- sufixos BHP/BHN viram códigos + e − (banco de horas); CEP/CP/CRO viram o código do serviço.
"""
import datetime as dt
import os
import re
import unicodedata

import openpyxl

DIR_FONTES = os.path.expanduser("~/Downloads/fwdarquivosdaescala")
nfc = lambda s: unicodedata.normalize("NFC", s)

ARQUIVOS = {
    # mês: (arquivo, aba, coluna da Manhã ou None = achar pelo cabeçalho, tem Noitinha?)
    # Cada arquivo carrega a semana de virada do mês vizinho (abril 1–3 está no fim
    # do arquivo de março, julho 1–5 no de junho…): o leitor segue a numeração e
    # troca o mês quando o número do dia volta pra trás.
    1: ("ESCALA FINAL JANEIRO 2026.xlsx", "Table 1", None, True),
    2: ("Escala fevereiro UTI HCB 2026.xlsx", "Planilha1", 4, False),
    3: ("Escala UTI - março 2026.xlsx", "Table 1", 4, False),
    4: ("Escala abril 2026 UTI HCB.xlsx", "Table 1", 4, True),
    5: ("Escala maio UTI - HCB - última revisão 09.04.26.docx.xlsx", "Table 1", None, True),
    6: ("Escala junho - UTI HCB.xlsx", "Junho 2026", None, True),
    7: ("Escala juLHO - UTI HCB - apos correcao de carga horaria.xlsx", "Julho 2026", 4, True),
    8: ("Escala agosto - UTI HCB.xlsx", "Agosto 2026", 4, True),
    9: ("Escala setembro - UTI HCB.xlsx", "Setembro 2026", 4, True),
}

DIAS_SEMANA = {"SEGUNDA": 0, "TERÇA": 1, "TERCA": 1, "QUARTA": 2, "QUINTA": 3,
               "SEXTA": 4, "SÁBADO": 5, "SABADO": 5, "DOMINGO": 6}

# sufixos de anotação que vêm colados no nome. BHP/BHN/CEP/CP/CRO são LIDOS
# (viram código: M+ M- CEP CP CRO); Pr/PR/ICDF/(…) são só descartados
ANOTACAO = re.compile(r"\b(BHP|BHN|CEP|CRO|CP)\b", re.I)
SUFIXOS = re.compile(r"\s+(BHP|BHN|BH|CRO|CP|CEP|Pr|PR|ICDF|\(.*\))\s*$", re.I)
DISPENSA = re.compile(r"\bBHN\b", re.I)
# linhas que não são pessoa
NAO_PESSOA = re.compile(r"^(niver|anivers|obs\b|feriado|total|legenda|coord)|^(manh[ãa]|tarde|noite|noitinha|dia)$", re.I)

# a grade escreve alguns nomes diferente do roster
ALIASES = {"Ste": "Stephanie", "Patricia Abreu": "Patricia", "PatiAbreu": "Patricia",
           "Msalomao": "MSalomão", "Mpinheiro": "MPinheiro", "Lelemos": "LeLemos",
           "Pjamile": "Pjamile", "Isabela": "IsaRibeiro", "Marina": "MSalomão"}


def _anotacao(nome):
    """BHP / BHN / CEP / CP / CRO colado no nome, ou None."""
    m = ANOTACAO.search(nfc(str(nome)))
    return m.group(1).upper() if m else None


def _limpar(nome):
    s = nfc(str(nome).strip())
    if not s or NAO_PESSOA.match(s):
        return None
    anterior = None
    while anterior != s:
        anterior = s
        s = SUFIXOS.sub("", s).strip()
    return s or None


def _canonico(nome, roster):
    nome = ALIASES.get(nome, nome)
    """casa o nome da grade com o apelido do roster, ignorando caixa e acento."""
    chave = unicodedata.normalize("NFD", nome.lower()).encode("ascii", "ignore")
    for apelido in roster:
        alvo = unicodedata.normalize("NFD", apelido.lower()).encode("ascii", "ignore")
        if chave == alvo:
            return apelido
    return None


def _dia_da_celula(v, mes_atual):
    """(dia, mês) a partir da célula do dia: inteiro, texto '6', '02/08' ou a data
    que o Excel inventou (1900-01-06 = dia 6). None se não é um dia."""
    if isinstance(v, (dt.datetime, dt.date)):
        return (v.day, mes_atual) if v.year == 1900 else (v.day, v.month)
    if isinstance(v, (int, float)) and float(v).is_integer() and 1 <= int(v) <= 31:
        return (int(v), mes_atual)
    if isinstance(v, str):
        t = v.strip()
        if t.isdigit() and 1 <= int(t) <= 31:
            return (int(t), mes_atual)
        m = re.match(r"^(\d{1,2})/(\d{1,2})$", t)
        if m:
            return (int(m.group(1)), int(m.group(2)))
    return None


def _ultimo(mes):
    return (dt.date(2026, mes % 12 + 1, 1) - dt.timedelta(days=1)).day if mes < 12 else 31


def ler_grade(mes, roster):
    nome_arq, aba, cM, tem_nt = ARQUIVOS[mes]
    wb = openpyxl.load_workbook(os.path.join(DIR_FONTES, nome_arq), data_only=True)
    ws = wb[aba]

    # coluna da Manhã: explícita, ou pelo cabeçalho; janeiro não tem cabeçalho → 3
    for r in range(1, (min(ws.max_row, 60) + 1) if cM is None else 1):
        for c in range(1, min(ws.max_column, 12) + 1):
            v = ws.cell(row=r, column=c).value
            if v and nfc(str(v).strip()).upper().startswith("MANH"):
                cM = c
                break
        if cM:
            break
    if cM is None:
        cM = 3
    colunas = [(cM, "M"), (cM + 1, "T"), (cM + 2, "N")] + ([(cM + 3, "NT")] if tem_nt else [])

    turnos = {}           # (apelido, date) -> {janela: anotação}
    mes_atual, dia_atual = mes, None
    naocasados = set()
    for r in range(1, ws.max_row + 1):
        linha = [ws.cell(row=r, column=c).value for c in range(1, cM)]
        achado = next((d for d in (_dia_da_celula(v, mes_atual) for v in linha) if d), None)
        nome_dow = next((DIAS_SEMANA[nfc(str(v)).strip().upper()] for v in linha
                         if isinstance(v, str)
                         and nfc(str(v)).strip().upper() in DIAS_SEMANA), None)
        if achado:
            numero, m2 = achado
            if m2 != mes_atual:
                mes_atual = m2                      # '02/08' num arquivo de julho
            elif dia_atual is not None and numero < dia_atual:
                mes_atual = mes_atual % 12 + 1      # a numeração voltou: mês seguinte
            if numero > _ultimo(mes_atual):
                continue
            dia_atual = numero
        elif nome_dow is not None and dia_atual is not None:
            # avança do dia corrente até casar o dia-da-semana (junho perde o número)
            candidato = dia_atual + 1
            while candidato <= _ultimo(mes_atual) and \
                    dt.date(2026, mes_atual, candidato).weekday() != nome_dow:
                candidato += 1
            if candidato <= _ultimo(mes_atual):
                dia_atual = candidato
            else:
                mes_atual = mes_atual % 12 + 1
                dia_atual = 1
                while dt.date(2026, mes_atual, dia_atual).weekday() != nome_dow:
                    dia_atual += 1
        if dia_atual is None:
            continue
        data = dt.date(2026, mes_atual, dia_atual)
        for coluna, janela in colunas:
            v = ws.cell(row=r, column=coluna).value
            if v in (None, ""):
                continue
            limpo = _limpar(v)
            if not limpo:
                continue
            apelido = _canonico(limpo, roster)
            if apelido is None:
                naocasados.add(limpo)
                continue
            turnos.setdefault((apelido, data), {})[janela] = _anotacao(v)

    dias = {}
    for (apelido, data), janelas in turnos.items():
        dias.setdefault(data, {})[apelido] = (_compor(janelas), "grade")
    return dias, sorted(naocasados)


def _compor(janelas):
    """{janela: anotação} de um dia → código da planilha.

    Manhã + tarde = D. BHP num período vira '+', BHN vira '-': M+ (manhã a mais),
    Dm+ (dia em que a manhã é BHP), M- (dispensa da manhã, não trabalha), Tm-
    (trabalha a tarde, a manhã é dispensa) etc. CEP/CP/CRO são serviços próprios.
    """
    for servico in ("CEP", "CP", "CRO"):
        if servico in janelas.values():
            return servico
    m, t = janelas.get("M", "ausente"), janelas.get("T", "ausente")
    tem_m, tem_t = "M" in janelas, "T" in janelas
    if tem_m and tem_t:
        if m == "BHN" and t == "BHN":
            return "D-"
        if m == "BHN":
            return "Tm-"
        if t == "BHN":
            return "Mt-"
        if m == "BHP" and t == "BHP":
            return "D+"
        if m == "BHP":
            return "Dm+"
        if t == "BHP":
            return "Dt+"
        return "D"           # dia + noite no mesmo dia: 18h — o validador pega
    if tem_m:
        return {"BHP": "M+", "BHN": "M-"}.get(m, "M")
    if tem_t:
        return {"BHP": "T+", "BHN": "T-"}.get(t, "T")
    if "N" in janelas:
        return {"BHP": "N+", "BHN": "N-"}.get(janelas["N"], "N")
    return "NT"


def importar(roster):
    """DIAS de todas as grades. Quando dois arquivos trazem a mesma data (a semana
    de virada), manda o arquivo cujo mês nominal é o mês da data."""
    DIAS, dono, relatorio = {}, {}, {}
    for mes in sorted(ARQUIVOS):
        dias, naocasados = ler_grade(mes, roster)
        for data, pessoas in dias.items():
            if data in DIAS and dono[data] == data.month:
                continue                          # o arquivo do mês certo já escreveu
            DIAS[data] = dict(pessoas)
            dono[data] = mes
        relatorio[mes] = {"dias": len(dias), "naocasados": naocasados,
                          "lancamentos": sum(len(v) for v in dias.values()),
                          "datas": (min(dias), max(dias)) if dias else None}
    return DIAS, relatorio


if __name__ == "__main__":
    from senior_import import MAPA_NOMES
    roster = sorted(set(MAPA_NOMES.values()))
    DIAS, rel = importar(roster)
    for mes, r in rel.items():
        print(f"mês {mes}: {r['dias']} dias · {r['lancamentos']} lançamentos")
        if r["naocasados"]:
            print(f"   não casados: {r['naocasados']}")
    # sanidade: lotação por turno num dia útil
    from collections import Counter
    print("\n=== lotação reconstruída (mínimo útil: 14 M / 10 T / 7 N) ===")
    for data in sorted(DIAS)[:1] + [d for d in sorted(DIAS) if d.day in (2, 3, 10, 17)][:6]:
        c = Counter(v[0] for v in DIAS[data].values())
        wd = ["seg","ter","qua","qui","sex","sáb","dom"][data.weekday()]
        m = c["M"] + c["D"]; t = c["T"] + c["D"]; n = c["N"]
        print(f"  {data:%d/%m} ({wd})  manhã {m:2d} · tarde {t:2d} · noite {n:2d}")
