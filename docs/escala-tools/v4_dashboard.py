#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Aba DASHBOARD (apresentação) e DADOS DASH (agregados por fórmula).

Decisões de forma, seguindo a disciplina de visualização:
- número herói + fila de KPIs em vez de gráfico de uma barra;
- "alertas de 18h por mês" é ÊNFASE (uma série, um matiz) — a história é a queda
  depois de abril, não a comparação entre categorias;
- manhã/tarde/noite é ORDEM DO DIA, não identidade → rampa sequencial
  clara→escura, que também dispensa a briga de separação entre matizes;
- saldo contra a meta é POLARIDADE → par divergente coral↔azul (validado: todos
  os checks passam, protan ΔE 12,3; coral↔sage reprovava em protanopia);
- status sempre com rótulo, nunca cor sozinha.

Paleta validada por script, não escolhida no olho. Superfície creme #FFFAF3.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- identidade
BG, BG_ALT = "FFFAF3", "FAF3E8"
INK, INK2, INK3 = "3A2E2A", "6B5C56", "9A8A82"
LINE, LINE2 = "EBE8E5", "DAD3CD"
LAV, LAVI, LAVS = "A299CB", "5A4E8C", "ECEAF4"
AQUAS, SANDS, CORALS, SAGES = "E8F6F8", "FBF1E1", "FBE9E5", "ECF6E7"
# rampa sequencial: manhã → tarde → noite (monotônica, todas ≥3:1 no creme)
RAMPA = ["8D80C0", "6A57A8", "453A73"]
# par divergente: abaixo da meta ↔ acima da meta
NEG, POS = "C77264", "1A79A8"
# status (semântica do Colo)
OK, WARN, ERR = "5A6E50", "D9A85A", "C77264"
DISPLAY, CORPO, MAO = "Fraunces", "Nunito", "Caveat"

MESES = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
# geometria das abas mensais (idêntica em todas)
R_P0, R_ULT = 8, 73
R_COB, R_FALTA = 75, 78          # 75-77 cobertura · 78-80 falta
# +6: as colunas C..H viraram as vésperas (fim do mês anterior que fecha a semana)
C_D1, C_DN = "I", "AM"        # cálculo continua nos 31 dias do mês
# AN..AS = virada (dias do mês seguinte até o domingo que fecha a última semana)
# AT..BE = Sem 1, BH 1, … Sem 6, BH 6 (as semanas vêm primeiro — V3, 01/09/26)
C_SEM = ["AT", "AV", "AX", "AZ", "BB", "BD"]
C_BH = ["AU", "AW", "AY", "BA", "BC", "BE"]
C_CH, C_FDS, C_SXN, C_FER = "BF", "BG", "BH", "BI"
C_META, C_SALDO, C_18H, C_NT = "BJ", "BK", "BL", "BM"

FINO = Side(style="thin", color=LINE)


def _sem_grade(ws):
    ws.sheet_view.showGridLines = False


def _txt(ws, cel, valor, fonte=CORPO, tam=10, cor=INK, negrito=False,
         italico=False, alin="left", vert="center", wrap=False):
    c = ws[cel] if isinstance(cel, str) else cel
    c.value = valor
    c.font = Font(name=fonte, size=tam, bold=negrito, italic=italico, color=cor)
    c.alignment = Alignment(horizontal=alin, vertical=vert, wrap_text=wrap)
    return c


def _bloco(ws, r1, c1, r2, c2, fundo=None, borda=True):
    """pinta um retângulo como se fosse um cartão."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cel = ws.cell(row=r, column=c)
            if fundo:
                cel.fill = PatternFill("solid", fgColor=fundo)
            if borda:
                cel.border = Border(
                    left=FINO if c == c1 else None, right=FINO if c == c2 else None,
                    top=FINO if r == r1 else None, bottom=FINO if r == r2 else None)
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r1, end_column=c2)


def _tem_dado(mes):
    """o mês foi montado? mês vazio não tem 'buraco' — tem ausência de escala."""
    return f"COUNTA({mes}!${C_D1}${R_P0}:${C_DN}${R_ULT})>0"


def _dias_com_gente(mes):
    """máscara 1/0 por dia: só dia que tem alguém lançado conta como dia da
    escala (novembro com só o 01/11 lançado não pode virar '852 buracos')."""
    return (f"(({mes}!${C_D1}${R_COB}:${C_DN}${R_COB}"
            f"+{mes}!${C_D1}${R_COB+1}:${C_DN}${R_COB+1}"
            f"+{mes}!${C_D1}${R_COB+2}:${C_DN}${R_COB+2})>0)")


def _falta_total(mes, guardado=True):
    """soma dos três turnos que faltam para o mínimo, só nos dias montados."""
    soma = "+".join(f"SUMPRODUCT({_dias_com_gente(mes)}*{mes}!${C_D1}${R_FALTA+i}:${C_DN}${R_FALTA+i})"
                    for i in range(3))
    return f'IF({_tem_dado(mes)},{soma},"")' if guardado else soma


def _dias_completos(mes, guardado=True):
    """dias montados em que os três turnos bateram o mínimo (falta zero nos três)."""
    if guardado:
        return f'IF({_tem_dado(mes)},{_dias_completos(mes, False)},"")'
    return (f"SUMPRODUCT({_dias_com_gente(mes)}*({mes}!${C_D1}${R_FALTA}:${C_DN}${R_FALTA}"
            f"+{mes}!${C_D1}${R_FALTA+1}:${C_DN}${R_FALTA+1}"
            f"+{mes}!${C_D1}${R_FALTA+2}:${C_DN}${R_FALTA+2}=0)"
            f"*({mes}!${C_D1}${R_FALTA}:${C_DN}${R_FALTA}<>\"\"))")


def aba_dados(wb):
    """agregados por mês — fonte dos gráficos. Fica oculta."""
    ws = wb.create_sheet("DADOS DASH")
    _sem_grade(ws)
    _txt(ws, "A1", "Agregados por mês · Fonte dos gráficos do painel",
         tam=11, negrito=True, cor=LAVI)
    _txt(ws, "A2", "Aba de apoio: não digitar nada aqui, tudo é fórmula das abas mensais",
         tam=9, italico=True, cor=INK3)
    cabec = ["Mês", "Alertas 18h", "Buracos", "Dias completos", "Manhã",
             "Tarde", "Noite", "Noites", "Horas noturnas", "Fora da meta"]
    for i, h in enumerate(cabec, start=1):
        c = _txt(ws, ws.cell(row=4, column=i), h, tam=9, negrito=True, cor="FFFFFF",
                 alin="center", wrap=True)
        c.fill = PatternFill("solid", fgColor=LAVI)
        ws.column_dimensions[get_column_letter(i)].width = 11 if i > 1 else 7
    for k, mes in enumerate(MESES):
        r = 5 + k
        _txt(ws, ws.cell(row=r, column=1), mes, tam=9, negrito=True, cor=INK)
        formulas = [
            f"=SUM({mes}!${C_18H}${R_P0}:${C_18H}${R_ULT})",
            f"={_falta_total(mes)}",
            f"={_dias_completos(mes)}",
            # lotação média por turno nos dias que existem
            f'=IF({_tem_dado(mes)},IFERROR(AVERAGEIF({mes}!${C_D1}${R_COB}:${C_DN}${R_COB},">0"),0),"")',
            f'=IF({_tem_dado(mes)},IFERROR(AVERAGEIF({mes}!${C_D1}${R_COB+1}:${C_DN}${R_COB+1},">0"),0),"")',
            f'=IF({_tem_dado(mes)},IFERROR(AVERAGEIF({mes}!${C_D1}${R_COB+2}:${C_DN}${R_COB+2},">0"),0),"")',
            f'=COUNTIF({mes}!${C_D1}${R_P0}:${C_DN}${R_ULT},"N")'
            f'+COUNTIF({mes}!${C_D1}${R_P0}:${C_DN}${R_ULT},"N+")',
            # noite 19–07h cobre integralmente as 7h da janela 22h–05h
            f'=(COUNTIF({mes}!${C_D1}${R_P0}:${C_DN}${R_ULT},"N")'
            f'+COUNTIF({mes}!${C_D1}${R_P0}:${C_DN}${R_ULT},"N+"))*7'
            f'+COUNTIF({mes}!${C_D1}${R_P0}:${C_DN}${R_ULT},"NT")*3',
            f'=SUMPRODUCT(--(ABS({mes}!${C_SALDO}${R_P0}:${C_SALDO}${R_ULT})>12))',
        ]
        # mês sem escala fica vazio em TODAS as colunas: zero e "não existe ainda"
        # são coisas diferentes, e misturar as duas foi o que fez novembro
        # aparecer com "871 buracos" na primeira versão
        for i, f in enumerate(formulas, start=2):
            if not f.startswith(f"=IF({_tem_dado(mes)}"):
                f = f'=IF({_tem_dado(mes)},{f[1:]},"")'
            cel = ws.cell(row=r, column=i, value=f)
            cel.font = Font(name=CORPO, size=9, color=INK2)
            cel.alignment = Alignment(horizontal="center")
            cel.number_format = "0.0" if i in (5, 6, 7) else "0"

    # distribuição do saldo no mês vivo — o trabalho aqui é POLARIDADE
    # (abaixo/acima da meta), então as faixas são divergentes em volta do zero
    _txt(ws, "A20", "Distribuição do saldo · Outubro", tam=10, negrito=True, cor=LAVI)
    _txt(ws, ws.cell(row=21, column=1), "Faixa", tam=9, negrito=True, cor="FFFFFF").fill = \
        PatternFill("solid", fgColor=LAVI)
    for j, rot in enumerate(("Devendo", "A mais")):
        c = _txt(ws, ws.cell(row=21, column=2 + j), rot, tam=9, negrito=True,
                 cor="FFFFFF", alin="center")
        c.fill = PatternFill("solid", fgColor=LAVI)
    faixa = f"OUT!${C_SALDO}${R_P0}:${C_SALDO}${R_ULT}"
    # rótulo curto: o eixo de um gráfico de barras corta texto longo
    # rótulo com SINAL: "até 12h" apareceria duas vezes e viraria ambiguidade
    LINHAS = [
        ("−24h ou mais", f'=COUNTIFS({faixa},"<-24")', None),
        ("−12 a −24h",   f'=COUNTIFS({faixa},">=-24",{faixa},"<-12")', None),
        ("−1 a −12h",    f'=COUNTIFS({faixa},">=-12",{faixa},"<0")', None),
        ("na meta",      f'=COUNTIFS({faixa},"=0")', None),
        ("+1 a +12h",    None, f'=COUNTIFS({faixa},">0",{faixa},"<=12")'),
        ("+12 a +24h",   None, f'=COUNTIFS({faixa},">12",{faixa},"<=24")'),
        ("+24h ou mais", None, f'=COUNTIFS({faixa},">24")'),
    ]
    for i, (rot, devendo, amais) in enumerate(LINHAS):
        r = 22 + i
        _txt(ws, ws.cell(row=r, column=1), rot, tam=9, cor=INK2)
        for j, f in enumerate((devendo, amais)):
            if f is None:
                continue
            cel = ws.cell(row=r, column=2 + j, value=f)
            cel.font = Font(name=CORPO, size=9, color=INK)
            cel.alignment = Alignment(horizontal="center")
    return ws


def _tip(cel, chave):
    import v4_dados as _D
    from openpyxl.comments import Comment
    texto = _D.TOOLTIPS.get(chave)
    if texto:
        cm = Comment(texto, "colo ritmo")
        cm.width, cm.height = 320, 130
        cel.comment = cm
    return cel


TIP_TILE = {"buracos de cobertura no mês": "buracos dash",
            "dias completos": "dias completos dash",
            "alertas de 18h": "alertas 18h dash",
            "horas noturnas": "noturnas dash"}


def _tile(ws, r1, c1, r2, c2, rotulo, formula, formato="0", nota="",
          cor_num=LAVI, fundo=BG_ALT, tam_num=26):
    """cartão de KPI: rótulo pequeno em cima, número grande, nota embaixo."""
    _bloco(ws, r1, c1, r2, c2, fundo=fundo)
    _tip(_txt(ws, ws.cell(row=r1, column=c1), rotulo.upper(), tam=8, negrito=True,
              cor=INK3, alin="center"), TIP_TILE.get(rotulo.lower(), ""))
    ws.merge_cells(start_row=r1 + 1, start_column=c1, end_row=r1 + 2, end_column=c2)
    n = ws.cell(row=r1 + 1, column=c1, value=formula)
    n.font = Font(name=DISPLAY, size=tam_num, bold=True, color=cor_num)
    n.alignment = Alignment(horizontal="center", vertical="center")
    n.number_format = formato
    ws.merge_cells(start_row=r2, start_column=c1, end_row=r2, end_column=c2)
    _txt(ws, ws.cell(row=r2, column=c1), nota, fonte=CORPO, tam=8, cor=INK3,
         alin="center", wrap=True)


def aba_dashboard(wb, mes_vivo="OUT", nome_mes="outubro"):
    ws = wb.create_sheet("DASHBOARD", 0)
    ws.sheet_properties.tabColor = LAVI
    _sem_grade(ws)
    larguras = [3, 14, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 3]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r, h in ((1, 8), (2, 22), (3, 30), (4, 10), (5, 16)):
        ws.row_dimensions[r].height = h

    # ---------------- cabeçalho
    _txt(ws, "B2", "Colo Ritmo · Hospital da Criança de Brasília",
         fonte=MAO, tam=13, cor=LAV)
    ws.merge_cells("B3:H3")
    _txt(ws, "B3", f"Painel da escala · {nome_mes.capitalize()} de 2026",
         fonte=DISPLAY, tam=22, negrito=True, cor=INK)
    ws.merge_cells("B5:H5")
    _txt(ws, "B5", "Toda contagem aqui é fórmula das abas mensais — digitou um código na "
                   "escala, este painel se refaz sozinho.", tam=9, italico=True, cor=INK3)
    _bloco(ws, 2, 2, 2, 14, fundo=None, borda=False)
    for c in range(2, 15):
        ws.cell(row=6, column=c).border = Border(bottom=Side(style="thin", color=LINE2))

    # ---------------- fila de KPIs · o mês que vai publicar
    ws.row_dimensions[8].height = 20
    _txt(ws, "B8", f"O mês que vai publicar · {nome_mes.capitalize()}", fonte=DISPLAY, tam=12,
         negrito=True, cor=LAVI)
    for r, h in ((9, 14), (10, 20), (11, 20), (12, 36)):
        ws.row_dimensions[r].height = h
    M = mes_vivo
    # herói: o número que decide se a escala pode ser publicada
    _tile(ws, 9, 2, 12, 4, "buracos de cobertura no mês",
          f"={_falta_total(M)}", nota="soma de tudo que falta pro mínimo, "
          "somando os três turnos e os 31 dias", cor_num=LAVI, fundo=LAVS, tam_num=40)
    _tile(ws, 9, 5, 12, 6, "dias completos",
          f"={_dias_completos(M)}&\"/\"&SUMPRODUCT({_dias_com_gente(M)}*1)",
          formato="General", nota="dias em que os três turnos bateram o mínimo")
    _tile(ws, 9, 7, 12, 8, "alertas de 18h",
          f"=SUM({M}!${C_18H}${R_P0}:${C_18H}${R_ULT})",
          nota="o alvo é zero · art. 66 CLT", cor_num=ERR, fundo=CORALS)
    _tile(ws, 9, 9, 12, 10, "fora da meta",
          f"=SUMPRODUCT(--(ABS({M}!${C_SALDO}${R_P0}:${C_SALDO}${R_ULT})>12))",
          nota="pessoas com mais de 12h de desvio, pra cima ou pra baixo")
    # semanas com BH ≠ 0: o que ainda não fechou (ou está no banco) em outubro
    fora_bh = "+".join(
        f'SUMPRODUCT(({M}!${c}${R_P0}:${c}${R_ULT}<>"")*({M}!${c}${R_P0}:${c}${R_ULT}<>0))'
        for c in C_BH)
    _tile(ws, 9, 11, 12, 12, "semanas com BH ≠ 0",
          f"={fora_bh}",
          nota="pessoa-semanas fora do alvo · BHP (+) ou BHN/faltou (−)", fundo=SANDS)
    _tile(ws, 9, 13, 12, 14, "horas noturnas",
          f'=(COUNTIF({M}!${C_D1}${R_P0}:${C_DN}${R_ULT},"N")'
          f'+COUNTIF({M}!${C_D1}${R_P0}:${C_DN}${R_ULT},"N+"))*7'
          f'+COUNTIF({M}!${C_D1}${R_P0}:${C_DN}${R_ULT},"NT")*3',
          nota="janela 22h–05h · insumo do adicional noturno", fundo=AQUAS)

    # nota: a fonte de outubro
    ws.row_dimensions[13].height = 26
    av = _txt(ws, "B13", "Outubro é a versão da Mari (Sheet vivo, 01/09) com as correções "
              "da checagem dela e do Marcos — cada mudança está na aba CHECAGEM OUT. "
              "Setembro vem da grade do grupo. Nada aqui é proposta do gerador.",
              tam=9, italico=True, cor=LAVI, wrap=True)
    ws.merge_cells("B13:N13")
    for c in range(2, 15):
        ws.cell(row=13, column=c).fill = PatternFill("solid", fgColor=LAVS)

    # ---------------- o ano
    ws.row_dimensions[14].height = 24
    _txt(ws, "B14", "O ano · Para comparar", fonte=DISPLAY, tam=12, negrito=True, cor=LAVI)
    _txt(ws, "B15", "As regras duras apertaram junto com a mudança de abril, e funcionou: "
                    "os alertas de 18h caíram de ~15 por mês para ~2.",
         tam=9, italico=True, cor=INK2)
    ws.merge_cells("B15:J15")

    # tabela do ano (também é a "table view" dos gráficos)
    r0 = 17
    cabec = ["Mês", "Alertas 18h", "Buracos", "Dias completos",
             "Manhã", "Tarde", "Noite", "Horas noturnas"]
    tip_tab = {"Alertas 18h": "alertas 18h dash", "Buracos": "buracos dash",
               "Dias completos": "dias completos dash", "Manhã": "lotação dash",
               "Tarde": "lotação dash", "Noite": "lotação dash",
               "Horas noturnas": "noturnas dash"}
    for i, h in enumerate(cabec):
        c = _txt(ws, ws.cell(row=r0, column=2 + i), h, tam=8, negrito=True,
                 cor="FFFFFF", alin="center", wrap=True)
        c.fill = PatternFill("solid", fgColor=LAVI)
        _tip(c, tip_tab.get(h, ""))
    for k, mes in enumerate(MESES):
        r = r0 + 1 + k
        _txt(ws, ws.cell(row=r, column=2), mes.capitalize(), tam=9, negrito=True, cor=INK)
        for i, col_dados in enumerate("BCDEFGI"):
            cel = ws.cell(row=r, column=3 + i,
                          value=f"='DADOS DASH'!{col_dados}{5+k}")
            cel.font = Font(name=CORPO, size=9, color=INK2)
            cel.alignment = Alignment(horizontal="center")
            cel.number_format = "0.0" if col_dados in "EFG" else "0"
            cel.border = Border(bottom=Side(style="thin", color=LINE))
        # faixa alternada, quente
        if k % 2 == 1:
            for c in range(2, 10):
                if not ws.cell(row=r, column=c).fill.fgColor.rgb or \
                        ws.cell(row=r, column=c).fill.fgColor.rgb == "00000000":
                    ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BG_ALT)
    ws.freeze_panes = "B9"
    return ws, r0
